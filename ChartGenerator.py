import matplotlib.pyplot as plt
import numpy as np
import io
import tempfile
import os
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas

class ChartGenerator:
    """
    Classe pour générer différents types de graphiques avec matplotlib
    et les exporter dans différents formats.
    """
    
    @staticmethod
    def create_line_chart(data, title="Graphique en courbes", xlabel="X", ylabel="Y"):
        """
        Crée un graphique en courbes à partir des données fournies.
        
        Args:
            data (dict): Dictionnaire contenant 'x' et 'y', ou seulement 'y'.
            title (str): Titre du graphique.
            xlabel (str): Étiquette de l'axe X.
            ylabel (str): Étiquette de l'axe Y.
            
        Returns:
            matplotlib.figure.Figure: L'objet figure contenant le graphique.
        """
        fig, ax = plt.subplots(figsize=(10, 6))
        
        if 'x' in data and 'y' in data:
            # Si x et y sont fournis
            if isinstance(data['y'][0], list) or isinstance(data['y'][0], np.ndarray):
                # Multiples séries
                for i, y_series in enumerate(data['y']):
                    ax.plot(data['x'], y_series, marker='o', label=f'Série {i+1}')
            else:
                # Série unique
                ax.plot(data['x'], data['y'], marker='o', label='Données')
        else:
            # Si seulement y est fourni, utiliser les indices comme x
            if isinstance(data['y'][0], list) or isinstance(data['y'][0], np.ndarray):
                # Multiples séries
                for i, y_series in enumerate(data['y']):
                    ax.plot(range(len(y_series)), y_series, marker='o', label=f'Série {i+1}')
            else:
                # Série unique
                ax.plot(range(len(data['y'])), data['y'], marker='o', label='Données')
        
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True)
        ax.legend()
        
        fig.tight_layout()
        return fig
    
    @staticmethod
    def create_bar_chart(data, title="Graphique à barres", xlabel="Catégories", ylabel="Valeurs"):
        """
        Crée un graphique à barres à partir des données fournies.
        
        Args:
            data (dict): Dictionnaire contenant 'x' (catégories) et 'y' (valeurs).
            title (str): Titre du graphique.
            xlabel (str): Étiquette de l'axe X.
            ylabel (str): Étiquette de l'axe Y.
            
        Returns:
            matplotlib.figure.Figure: L'objet figure contenant le graphique.
        """
        fig, ax = plt.subplots(figsize=(10, 6))
        
        if 'x' in data and 'y' in data:
            # Si x (catégories) et y (valeurs) sont fournis
            ax.bar(data['x'], data['y'])
        else:
            # Si seulement y est fourni, utiliser les indices comme catégories
            ax.bar(range(len(data['y'])), data['y'])
        
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True, axis='y')
        
        # Rotation des étiquettes si nécessaire
        if 'x' in data and isinstance(data['x'][0], str) and any(len(str(x)) > 3 for x in data['x']):
            plt.xticks(rotation=45, ha='right')
        
        fig.tight_layout()
        return fig
    
    @staticmethod
    def create_pie_chart(data, title="Graphique en camembert"):
        """
        Crée un graphique en camembert à partir des données fournies.
        
        Args:
            data (dict): Dictionnaire contenant 'labels' et 'values'.
            title (str): Titre du graphique.
            
        Returns:
            matplotlib.figure.Figure: L'objet figure contenant le graphique.
        """
        fig, ax = plt.subplots(figsize=(10, 8))
        
        # Vérifier que les données contiennent des étiquettes et des valeurs
        if 'labels' in data and 'values' in data:
            # S'assurer que les longueurs sont les mêmes
            min_len = min(len(data['labels']), len(data['values']))
            labels = data['labels'][:min_len]
            values = data['values'][:min_len]
            
            # Calculer les pourcentages pour l'affichage
            total = sum(values)
            percentages = [100 * val / total for val in values]
            
            # Créer des étiquettes avec pourcentages
            labels_with_pct = [f'{l} ({p:.1f}%)' for l, p in zip(labels, percentages)]
            
            # Créer le camembert
            wedges, texts, autotexts = ax.pie(
                values, 
                labels=labels_with_pct,
                autopct='%1.1f%%',
                startangle=90,
                shadow=True,
            )
            
            # Améliorer la lisibilité des étiquettes
            for text in texts:
                text.set_fontsize(9)
            for autotext in autotexts:
                autotext.set_fontsize(9)
                autotext.set_color('white')
        
        ax.set_title(title)
        ax.axis('equal')  # Pour que le camembert soit circulaire
        
        fig.tight_layout()
        return fig
    
    @staticmethod
    def create_scatter_plot(data, title="Nuage de points", xlabel="X", ylabel="Y"):
        """
        Crée un nuage de points à partir des données fournies.
        
        Args:
            data (dict): Dictionnaire contenant 'x' et 'y'.
            title (str): Titre du graphique.
            xlabel (str): Étiquette de l'axe X.
            ylabel (str): Étiquette de l'axe Y.
            
        Returns:
            matplotlib.figure.Figure: L'objet figure contenant le graphique.
        """
        fig, ax = plt.subplots(figsize=(10, 6))
        
        if 'x' in data and 'y' in data:
            # Tracer le nuage de points
            ax.scatter(data['x'], data['y'], alpha=0.7, s=50)
            
            # Ajouter une ligne de tendance si possible
            try:
                z = np.polyfit(data['x'], data['y'], 1)
                p = np.poly1d(z)
                x_line = np.linspace(min(data['x']), max(data['x']), 100)
                ax.plot(x_line, p(x_line), "r--", alpha=0.7, label=f'Tendance: y={z[0]:.2f}x+{z[1]:.2f}')
                ax.legend()
            except:
                pass  # Si erreur dans le calcul de la tendance, l'ignorer
        
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True)
        
        fig.tight_layout()
        return fig
    
    @staticmethod
    def create_histogram(data, title="Histogramme", xlabel="Valeurs", ylabel="Fréquence", bins=10):
        """
        Crée un histogramme à partir des données fournies.
        
        Args:
            data (dict): Dictionnaire contenant 'values'.
            title (str): Titre du graphique.
            xlabel (str): Étiquette de l'axe X.
            ylabel (str): Étiquette de l'axe Y.
            bins (int): Nombre de bacs pour l'histogramme.
            
        Returns:
            matplotlib.figure.Figure: L'objet figure contenant le graphique.
        """
        fig, ax = plt.subplots(figsize=(10, 6))
        
        if 'values' in data:
            # Créer l'histogramme
            n, bins, patches = ax.hist(data['values'], bins=bins, alpha=0.7, color='skyblue', edgecolor='black')
            
            # Ajouter une courbe de densité si possible
            try:
                import scipy.stats as stats
                density = stats.gaussian_kde(data['values'])
                x = np.linspace(min(data['values']), max(data['values']), 200)
                ax.plot(x, density(x) * len(data['values']) * (bins[1] - bins[0]), 'r-', alpha=0.7, label='Densité')
                ax.legend()
            except ImportError:
                pass  # Si scipy n'est pas disponible, ignorer la courbe de densité
        
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True)
        
        fig.tight_layout()
        return fig
    
    @staticmethod
    def figure_to_image(fig, format='png', dpi=100):
        """
        Convertit une figure matplotlib en image.
        
        Args:
            fig (matplotlib.figure.Figure): Figure à convertir.
            format (str): Format de l'image ('png', 'jpg', etc.).
            dpi (int): Résolution de l'image.
            
        Returns:
            io.BytesIO: Objet contenant les données de l'image.
        """
        buf = io.BytesIO()
        fig.savefig(buf, format=format, dpi=dpi)
        buf.seek(0)
        return buf
    
    @staticmethod
    def generate_sample_data():
        """
        Génère des données d'exemple pour différents types de graphiques.
        
        Returns:
            dict: Dictionnaire contenant des données d'exemple pour chaque type de graphique.
        """
        # Données pour graphique en courbes
        x_line = list(range(1, 11))
        y_line1 = [x * 1.5 + np.random.normal(0, 2) for x in x_line]
        y_line2 = [x * 0.8 + np.random.normal(0, 1) for x in x_line]
        y_line3 = [15 - x * 0.5 + np.random.normal(0, 1.5) for x in x_line]
        
        # Données pour graphique à barres
        categories = ['A', 'B', 'C', 'D', 'E', 'F']
        values = [25, 40, 30, 55, 15, 45]
        
        # Données pour graphique en camembert
        pie_labels = ['Segment 1', 'Segment 2', 'Segment 3', 'Segment 4', 'Segment 5']
        pie_values = [35, 25, 20, 10, 10]
        
        # Données pour nuage de points
        x_scatter = np.random.uniform(0, 10, 50)
        y_scatter = 3 * x_scatter + 2 + np.random.normal(0, 5, 50)
        
        # Données pour histogramme
        histogram_values = np.random.normal(50, 15, 200)
        
        # Organiser les données par type de graphique
        sample_data = {
            'line': {
                'single': {'x': x_line, 'y': y_line1},
                'multi': {'x': x_line, 'y': [y_line1, y_line2, y_line3]}
            },
            'bar': {'x': categories, 'y': values},
            'pie': {'labels': pie_labels, 'values': pie_values},
            'scatter': {'x': x_scatter.tolist(), 'y': y_scatter.tolist()},
            'histogram': {'values': histogram_values.tolist()}
        }
        
        return sample_data
    
    @staticmethod
    def create_excel_with_charts(charts_data, output_path):
        """
        Crée un fichier Excel contenant plusieurs feuilles avec des graphiques.
        
        Args:
            charts_data (list): Liste de tuples (figure, nom_feuille).
            output_path (str): Chemin où sauvegarder le fichier Excel.
            
        Returns:
            bool: True si réussi, False sinon.
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image
        except ImportError:
            print("Les modules pandas et openpyxl sont requis pour cette fonctionnalité.")
            return False
        
        # Créer un workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Données"
        
        # Ajouter des données d'exemple à la première feuille
        ws['A1'] = "Données d'exemple pour les graphiques"
        
        # Pour chaque graphique
        for i, (fig, chart_name) in enumerate(charts_data):
            # Sauvegarder le graphique comme image temporaire
            temp_dir = tempfile.mkdtemp()
            img_path = os.path.join(temp_dir, f"chart_{i}.png")
            fig.savefig(img_path, format='png', dpi=150)
            plt.close(fig)  # Fermer la figure pour libérer la mémoire
            
            # Créer une nouvelle feuille pour chaque graphique
            sheet_name = chart_name[:31]  # Limiter à 31 caractères (limite Excel)
            ws = wb.create_sheet(title=sheet_name)
            
            # Ajouter l'image à la feuille
            img = Image(img_path)
            # Redimensionner l'image si nécessaire
            img.width = 600
            img.height = 400
            ws.add_image(img, 'B2')
            
            # Ajouter un titre
            ws['B1'] = chart_name
        
        # Enregistrer le workbook
        wb.save(output_path)
        
        # Nettoyer les fichiers temporaires
        for i in range(len(charts_data)):
            try:
                os.remove(os.path.join(temp_dir, f"chart_{i}.png"))
                os.rmdir(temp_dir)
            except:
                pass
                
        return True
