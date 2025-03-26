import tkinter as tk
from tkinter import (
    scrolledtext,
    Label,
    Button,
    Entry,
    Menu,
    Toplevel,
    END,
    W,
    EW,
    messagebox,
    Scale,
    Radiobutton,
    Scrollbar,
    Text,
    filedialog,
    ttk,
    Frame,
    LEFT,
    SOLID,
    NE,
)
import time
import threading
import json
import requests
import logging
import os
import socket
import base64
from urllib.parse import urlparse
from functools import partial
import re
import pyperclip
from pygments import highlight
from pygments.lexers import get_lexer_by_name, guess_lexer
from pygments.formatters import get_formatter_by_name
import mimetypes
import urllib.parse
import importlib
import urllib3
import subprocess
import sys

# Imports pour l'OCR
try:
    import pytesseract
    import cv2
    import numpy as np
    from PIL import Image, ImageTk, ImageEnhance, ImageFilter
    HAS_OCR = True
except ImportError:
    HAS_OCR = False
    logging.warning("OCR non disponible. Installez pytesseract, opencv-python et Pillow pour activer la fonctionnalité.")

# Désactivation des avertissements liés à la vérification SSL pour Eden
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def filter_unicode_chars(text):
    """Supprime les caractères Unicode qui sont hors de la plage supportée par Tkinter/Tcl"""
    if isinstance(text, str):
        # Supprimer tous les caractères en dehors de la plage U+0000-U+FFFF
        return ''.join(char for char in text if ord(char) <= 0xFFFF)
    return text


class EdenClient:
    """Client pour l'API Eden"""
    def __init__(self, timeout, proxy_config=None):
        self.timeout = timeout
        self.current_request = None
        self.api_url = "https://edenportal-sfr.eden-genai.com/eden/ask"
        self.api_key = "eden-9b3e45fd08267f2b399f63e4e3f77ef7"
        self.geni_id = "b6742807-1efa-4416-8d8a-f72f3ad64e9b"
        # Stocker le dernier message et la dernière réponse pour maintenir la cohérence
        self.last_message = None
        self.last_response = None


        # Configurer le proxy selon les paramètres (code ajouté)
        self.proxies = {}
        if proxy_config and proxy_config.get("proxy_enabled", False):
            username = proxy_config.get("proxy_username", "")
            password = proxy_config.get("proxy_password", "")
            host = proxy_config.get("proxy_host", "")
            port = proxy_config.get("proxy_port", "")
            
            if host and port:
                # Si username et password sont fournis, les inclure dans l'URL du proxy
                if username and password:
                    # Encoder correctement le mot de passe
                    encoded_password = urllib.parse.quote(password)
                    proxy_url = f"http://{username}:{encoded_password}@{host}:{port}"
                else:
                    proxy_url = f"http://{host}:{port}"
                
                self.proxies = {
                    'http': proxy_url,
                    'https': proxy_url
                }
                
                logging.info(f"Proxy configuré pour Eden: {proxy_url}")
            else:
                logging.info("Proxy activé pour Eden mais hôte ou port manquant")
        else:
            # Utiliser les paramètres de proxy système si disponibles
            self.proxies = get_proxy_settings()
            if self.proxies:
                logging.info("Paramètres de proxy système détectés et activés pour Eden")
            else:
                logging.info("Aucun paramètre de proxy détecté pour Eden")


    def chat_with_ai(self, user_message, message_history=None, model=None):
        headers = {
            "accept": "application/json",
            "api-key": self.api_key,
            "Content-Type": "application/json"
        }

        # Construction du contexte à partir de l'historique des messages
        context = ""
        if message_history and len(message_history) > 0:
            # Log de débogage pour voir l'historique reçu
            logging.info(f"EdenClient: Reçu {len(message_history)} messages d'historique")
            
            for msg in message_history:
                role = msg.get("role", "")
                content = msg.get("content", "")
                if role == "user":
                    context += f"Utilisateur: {content}\n\n"
                elif role == "assistant":
                    context += f"Assistant: {content}\n\n"
        
      
        # Ajouter le contexte au prompt si disponible
        prompt_with_context = user_message
        if context:
            prompt_with_context = f"Historique de la conversation:\n{context}\n\nNouvelle demande: {user_message}\n\nImportant: Assurez-vous de bien prendre en compte les échanges précédents dans votre réponse."

        # Récupérer l'option de vérification internet depuis la configuration
        internet_verify = self.config.get("eden_internet_verify", False) if hasattr(self, "config") else False
        
        payload = {
            "geniId": self.geni_id,
            "prompt": prompt_with_context,
            "language": "fr",
            "usage": True,
            "internetSearch": internet_verify  # Ajouter l'option de vérification internet
        }

        # Enregistrer ce message comme étant le dernier message
        self.last_message = user_message

        try:

            self.current_request = requests.post(
                self.api_url,
                headers=headers,
                json=payload,
                verify=False,
                timeout=self.timeout,
                proxies=self.proxies  # Ajout des proxies ici
            )
            self.current_request.raise_for_status()
            response_data = self.current_request.json()
            self.current_request = None
            
            # Stocker cette réponse comme dernière réponse
            response = response_data["response"]
            self.last_response = response
            
            return response
        except requests.exceptions.Timeout:
            logging.error("Timeout lors de l'appel à l'API Eden")
            return "L'appel à l'API Eden a expiré. Veuillez réessayer."
        except requests.exceptions.RequestException as e:
            logging.error(f"Erreur de connexion à l'API Eden: {e}")
            return f"Erreur de connexion: {e}"
        except Exception as e:
            logging.error(f"Une erreur inattendue s'est produite avec l'API Eden: {e}")
            return f"Une erreur s'est produite: {e}"

    def cancel_request(self):
        if self.current_request:
            try:
                self.current_request.close()
                logging.info("Requête Eden annulée")
                return True
            except Exception as e:
                logging.error(f"Erreur lors de l'annulation de la requête Eden: {e}")
        return False

try:
    from PIL import Image, ImageTk
except ImportError:
    logging.warning("PIL non disponible. La prévisualisation des images sera désactivée.")
    Image = None
    ImageTk = None

try:
    import openpyxl
except ImportError:
    logging.warning("openpyxl non disponible. La lecture des fichiers Excel sera limitée.")
    openpyxl = None

try:
    import pandas as pd
    import re
    import tempfile
    from io import BytesIO
except ImportError:
    logging.warning("pandas non disponible. Le traitement avancé des fichiers Excel sera désactivé.")
    pd = None
    
try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
    import numpy as np
    import io
except ImportError:
    logging.warning("matplotlib non disponible. La création de graphiques sera désactivée.")
    plt = None

# Désactivation des avertissements liés à la vérification SSL pour Eden
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


class OCRHandler:
    """Classe pour gérer l'extraction de texte des images"""
    
    @staticmethod
    def extract_text_online(image_path, api_key='helloworld', lang='fra', proxy_config=None):

        """Extrait le texte d'une image en utilisant l'API OCR.space"""
        try:
            logging.info(f"Tentative d'OCR en ligne pour l'image: {image_path}")
            logging.info(f"Paramètres: api_key={api_key}, lang={lang}")
            
            # Vérifier que le fichier existe
            if not os.path.exists(image_path):
                logging.error(f"Fichier introuvable: {image_path}")
                return f"Erreur: Fichier image introuvable: {image_path}"
            
            # Préparer l'image
            with open(image_path, 'rb') as f:
                img_data = f.read()
            
            logging.info(f"Taille de l'image: {len(img_data)} octets")
            
            # Paramètres de l'API
            payload = {
                'apikey': api_key,
                'language': lang,
                'OCREngine': '2',
                'scale': 'true',
                'isTable': 'false',
                'detectOrientation': 'true',
                'filetype': os.path.splitext(image_path)[1][1:].lower()  # Extension sans le point
            }
            
            logging.info(f"Paramètres API: {payload}")
            
            # Utiliser le nom de fichier réel et le type MIME correct
            filename = os.path.basename(image_path)
            mime_type = mimetypes.guess_type(image_path)[0] or 'image/jpeg'
            
            files = {
                'file': (filename, img_data, mime_type)
            }
            
            logging.info(f"Envoi de la requête à OCR.space avec mime_type: {mime_type}")
            

            # Configuration du proxy pour la requête OCR
            proxies = {}
            if proxy_config and proxy_config.get("proxy_enabled", False):
                username = proxy_config.get("proxy_username", "")
                password = proxy_config.get("proxy_password", "")
                host = proxy_config.get("proxy_host", "")
                port = proxy_config.get("proxy_port", "")
            
                if host and port:
                    # Si username et password sont fournis, les inclure dans l'URL du proxy
                    if username and password:
                        # Encoder correctement le mot de passe
                        encoded_password = urllib.parse.quote(password)
                        proxy_url = f"http://{username}:{encoded_password}@{host}:{port}"
                    else:
                        proxy_url = f"http://{host}:{port}"
                
                    proxies = {
                        'http': proxy_url,
                        'https': proxy_url
                    }
                    logging.info(f"Proxy configuré pour OCR online: {proxy_url}")
                else:
                    logging.info("Proxy activé pour OCR online mais hôte ou port manquant")
            else:
                # Utiliser les paramètres de proxy système si disponibles
                try:
                    proxies = get_proxy_settings() if 'get_proxy_settings' in globals() else {}
                    if proxies:
                        logging.info("Paramètres de proxy système détectés et activés pour OCR")
                    else:
                        logging.info("Aucun paramètre de proxy détecté pour OCR")
                except Exception as e:
                    logging.error(f"Erreur lors de la récupération des paramètres proxy: {e}")


            # Appel à l'API
            response = requests.post(
                'https://api.ocr.space/parse/image',
                files=files,
                data=payload,
                proxies=proxies,  # Utilisation des proxies configurés
                verify=False 
            )
            
            logging.info(f"Statut de réponse: {response.status_code}")
            
            if response.status_code == 200:
                result = response.json()
                logging.info(f"Réponse API: {result}")
                
                # Vérifier la présence d'erreurs dans la réponse
                if result.get('IsErroredOnProcessing', False):
                    error_message = result.get('ErrorMessage', ['Erreur inconnue'])[0]
                    logging.error(f"Erreur API OCR.space: {error_message}")
                    return f"Erreur OCR.space: {error_message}"
                
                if result.get('ParsedResults'):
                    parsed_text = result['ParsedResults'][0]['ParsedText']
                    if parsed_text.strip():
                        return parsed_text
                    else:
                        logging.warning("Texte extrait vide")
                        return "Aucun texte n'a pu être extrait de l'image."
                else:
                    logging.warning("Aucun résultat d'analyse dans la réponse")
                    return "Aucun texte n'a pu être extrait de l'image."
            else:
                error_text = response.text
                logging.error(f"Erreur API ({response.status_code}): {error_text}")
                return f"Erreur lors de l'appel à l'API: Code {response.status_code}, Détails: {error_text[:200]}"
                
        except Exception as e:
            logging.error(f"Exception lors de l'extraction du texte en ligne: {e}", exc_info=True)
            return f"Erreur lors de l'extraction du texte: {str(e)}"
    
    @staticmethod
    def preprocess_image(image_path):
        """Prétraite l'image pour améliorer la qualité de l'OCR"""
        try:
            logging.info(f"Prétraitement de l'image: {image_path}")
            
            # Lire l'image avec PIL
            image = Image.open(image_path)
            
            # Sauvegarder la taille d'origine
            original_width, original_height = image.size
            logging.info(f"Taille d'origine: {original_width}x{original_height}")
            
            # Redimensionner l'image si elle est très grande, tout en gardant le ratio
            if original_width > 2000 or original_height > 2000:
                max_dim = 2000
                scale_factor = min(max_dim / original_width, max_dim / original_height)
                new_width = int(original_width * scale_factor)
                new_height = int(original_height * scale_factor)
                image = image.resize((new_width, new_height), Image.Resampling.LANCZOS)
                logging.info(f"Image redimensionnée à: {new_width}x{new_height}")
            
            # Convertir en RGB si nécessaire (pour les images avec transparence)
            if image.mode == 'RGBA':
                logging.info("Conversion RGBA → RGB")
                # Créer un fond blanc et coller l'image dessus pour gérer la transparence
                background = Image.new('RGB', image.size, (255, 255, 255))
                background.paste(image, mask=image.split()[3])  # 3 est le canal alpha
                image = background
            elif image.mode != 'RGB' and image.mode != 'L':
                logging.info(f"Conversion {image.mode} → RGB")
                image = image.convert('RGB')
            
            # Convertir en niveaux de gris
            gray = image.convert('L')
            
            # Application d'un filtre de netteté pour améliorer les contours
            sharpened = gray.filter(ImageFilter.SHARPEN)
            
            # Améliorer le contraste
            enhancer = ImageEnhance.Contrast(sharpened)
            enhanced = enhancer.enhance(2.0)
            
            # Réduire le bruit
            denoised = enhanced.filter(ImageFilter.MedianFilter(size=3))
            
            # Binarisation adaptative (simulation, car PIL n'a pas de binarisation adaptative intégrée)
            # Utilisation d'un seuil simple mais efficace
            threshold = 150  # Ajusté pour une meilleure perception des textes
            binary = denoised.point(lambda x: 0 if x < threshold else 255, '1')
            
            # Sauvegarde temporaire pour débogage si nécessaire
            # debug_path = os.path.join(os.path.dirname(image_path), 'debug_ocr.png')
            # binary.save(debug_path)
            # logging.info(f"Image prétraitée sauvegardée pour débogage: {debug_path}")
            
            return binary
            
        except Exception as e:
            logging.error(f"Erreur lors du prétraitement de l'image: {e}", exc_info=True)
            return None
    
    @staticmethod
    def extract_text(image_path, lang='fra+eng', use_online=True, proxy_config=None):
        """Extrait le texte d'une image avec OCR"""
        logging.info(f"Demande d'extraction de texte pour {image_path}, mode online: {use_online}")
        
        # Vérifier que le fichier existe
        if not os.path.exists(image_path):
            logging.error(f"Fichier introuvable: {image_path}")
            return f"Erreur: Fichier image introuvable: {image_path}"
            
        # Vérifier que c'est bien une image
        try:
            with Image.open(image_path) as img:
                format = img.format
                size = img.size
                logging.info(f"Image validée: format={format}, taille={size}")
        except Exception as e:
            logging.error(f"Fichier non reconnu comme image: {e}")
            return f"Erreur: Le fichier n'est pas reconnu comme une image valide: {str(e)}"
            
        if use_online:
            result = OCRHandler.extract_text_online(image_path, lang=lang.split('+')[0], proxy_config=proxy_config)
            # Si le résultat indique qu'aucun texte n'a été trouvé et que HAS_OCR est True,
            # essayer avec le mode local comme solution de secours
            if result == "Aucun texte n'a pu être extrait de l'image." and HAS_OCR:
                logging.info("OCR en ligne n'a pas trouvé de texte, essai avec OCR local")
                try:
                    local_result = OCRHandler._extract_text_local(image_path, lang)
                    if local_result and local_result != "Aucun texte n'a pu être extrait de l'image.":
                        return local_result
                except Exception as e:
                    logging.error(f"Erreur lors de la tentative d'OCR local (solution de secours): {e}")
            return result
            
        # Mode local
        return OCRHandler._extract_text_local(image_path, lang)
    
    @staticmethod
    def _extract_text_local(image_path, lang='fra+eng'):
        """Méthode interne pour l'extraction de texte en mode local avec Tesseract"""
        try:
            if not HAS_OCR:
                return "OCR local non disponible. Installez pytesseract, opencv-python et Pillow."
            
            logging.info(f"Tentative d'OCR local pour {image_path}")
            
            # Prétraiter l'image
            processed_image = OCRHandler.preprocess_image(image_path)
            if processed_image is None:
                return "Erreur lors du prétraitement de l'image."
            
            # Configuration de Tesseract
            custom_config = r'--oem 3 --psm 6'
            
            # Extraire le texte
            text = pytesseract.image_to_string(processed_image, lang=lang, config=custom_config)
            
            # Nettoyer le texte
            text = text.strip()
            
            if not text:
                logging.warning("OCR local n'a pas trouvé de texte")
                return "Aucun texte n'a pu être extrait de l'image."
                
            logging.info(f"OCR local a extrait {len(text)} caractères")
            return text
            
        except Exception as e:
            logging.error(f"Erreur lors de l'extraction de texte en local: {e}", exc_info=True)
            return f"Erreur lors de l'extraction du texte en local: {str(e)}"

# Intégration de ChartGenerator directement dans le fichier
class ChartGenerator:
    """Classe pour générer des graphiques et les exporter vers Excel"""
    
    @staticmethod
    def figure_to_image(fig):
        """Convertit une figure matplotlib en image"""
        buf = io.BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        return buf
    
    @staticmethod
    def create_line_chart(data, title='', xlabel='', ylabel=''):
        """Crée un graphique en ligne à partir des données"""
        fig, ax = plt.subplots(figsize=(8, 5))
        
        # Vérifier si les données contiennent des séries multiples
        if isinstance(data, dict) and 'series' in data:
            for series in data['series']:
                ax.plot(series['x'] if 'x' in series else range(len(series['y'])), 
                        series['y'], 
                        label=series.get('name', ''))
            ax.legend()
        else:
            # Données simples x,y
            if 'x' in data:
                ax.plot(data['x'], data['y'])
            else:
                ax.plot(data['y'])
        
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True)
        
        return fig
    
    @staticmethod
    def create_bar_chart(data, title='', xlabel='', ylabel=''):
        """Crée un graphique à barres à partir des données"""
        fig, ax = plt.subplots(figsize=(8, 5))
        
        if 'x' in data:
            ax.bar(data['x'], data['y'])
        else:
            ax.bar(range(len(data['y'])), data['y'])
        
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        
        # Si nous avons des étiquettes x, les faire pivoter pour une meilleure lisibilité
        if 'x' in data and isinstance(data['x'][0], str):
            plt.xticks(rotation=45, ha='right')
        
        ax.grid(True, axis='y')
        fig.tight_layout()
        
        return fig
    
    @staticmethod
    def create_pie_chart(data, title=''):
        """Crée un graphique en camembert à partir des données"""
        fig, ax = plt.subplots(figsize=(8, 5))
        
        ax.pie(data['values'], labels=data['labels'], autopct='%1.1f%%', 
               shadow=True, startangle=90)
        ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle
        ax.set_title(title)
        
        return fig
    
    @staticmethod
    def create_scatter_plot(data, title='', xlabel='', ylabel=''):
        """Crée un nuage de points à partir des données"""
        fig, ax = plt.subplots(figsize=(8, 5))
        
        if 'x' in data:
            ax.scatter(data['x'], data['y'])
        else:
            ax.scatter(range(len(data['y'])), data['y'])
        
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True)
        
        return fig
    
    @staticmethod
    def create_histogram(data, title='', xlabel='', ylabel='', bins=10):
        """Crée un histogramme à partir des données"""
        fig, ax = plt.subplots(figsize=(8, 5))
        
        ax.hist(data['values'], bins=bins)
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True)
        
        return fig
    
    @staticmethod
    def create_excel_with_charts(charts_data, output_file):
        """
        Crée un fichier Excel avec des graphiques
        
        Args:
            charts_data: Liste de tuples (figure, nom_feuille)
            output_file: Chemin du fichier Excel de sortie
        """
        if openpyxl is None:
            raise ImportError("openpyxl est requis pour créer des fichiers Excel")
        
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Pour chaque graphique
        for i, (fig, sheet_name) in enumerate(charts_data):
            # Créer une nouvelle feuille
            ws = wb.create_sheet(title=sheet_name[:31])  # Limité à 31 caractères
            
            # Enregistrer la figure dans un fichier temporaire
            img_data = ChartGenerator.figure_to_image(fig)
            
            # Créer l'image Excel
            img = XLImage(img_data)
            
            # Ajouter l'image à la feuille
            ws.add_image(img, 'B2')
            
            # Redimensionner les colonnes pour une meilleure visualisation
            ws.column_dimensions['A'].width = 5
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws.column_dimensions[col].width = 15
        
        # Enregistrer le classeur
        wb.save(output_file)
    
    @staticmethod
    def generate_sample_data():
        """Génère des données d'exemple pour différents types de graphiques"""
        import numpy as np
        
        # Données pour graphique linéaire
        x = np.linspace(0, 10, 100)
        y1 = np.sin(x)
        y2 = np.cos(x)
        
        line_data = {'x': x.tolist(), 'y': y1.tolist()}
        
        # Données multi-séries
        multi_line_data = {
            'series': [
                {'x': x.tolist(), 'y': y1.tolist(), 'name': 'Sinus'},
                {'x': x.tolist(), 'y': y2.tolist(), 'name': 'Cosinus'}
            ]
        }
        
        # Données pour graphique à barres
        bar_labels = ['A', 'B', 'C', 'D', 'E']
        bar_values = np.random.randint(1, 100, size=5).tolist()
        bar_data = {'x': bar_labels, 'y': bar_values}
        
        # Données pour camembert
        pie_labels = ['Groupe 1', 'Groupe 2', 'Groupe 3', 'Groupe 4']
        pie_values = np.random.randint(1, 100, size=4).tolist()
        pie_data = {'labels': pie_labels, 'values': pie_values}
        
        # Données pour nuage de points
        scatter_x = np.random.normal(0, 1, 50).tolist()
        scatter_y = np.random.normal(0, 1, 50).tolist()
        scatter_data = {'x': scatter_x, 'y': scatter_y}
        
        # Données pour histogramme
        histogram_values = np.random.normal(0, 1, 1000).tolist()
        histogram_data = {'values': histogram_values}
        
        return {
            'line': {
                'simple': line_data,
                'multi': multi_line_data
            },
            'bar': bar_data,
            'pie': pie_data,
            'scatter': scatter_data,
            'histogram': histogram_data
        }

# ------------------- Configuration -------------------


class Config:
    CONFIG_FILE = "config.json"
    HISTORY_FILE = "chat_history.json"
    PERSISTENCE_FILE = "application_state.json"

    DEFAULT_API_KEY = ""
    DEFAULT_API_URL = "https://openrouter.ai/api/v1/chat/completions"
    DEFAULT_MODEL = "google/gemini-2.0-flash-lite-preview-02-05:free"
    DEFAULT_TYPING_DELAY = 0.01
    DEFAULT_RESPONSE_MODE = "typing"
    DEFAULT_WINDOW_WIDTH = 800
    DEFAULT_WINDOW_HEIGHT = 600
    DEFAULT_TIMEOUT = 30
    DEFAULT_THEME = "light"
    DEFAULT_DUAL_CHAT = False  # Désactiver par défaut
    
    # Paramètres pour l'OCR
    DEFAULT_OCR_MODE = "online"  # "online" ou "local"
    DEFAULT_OCR_API_KEY = "helloworld"  # Clé par défaut pour OCR.space
    DEFAULT_OCR_LANG = "fra+eng"  # Langues par défaut
    DEFAULT_OCR_ENABLED = True  # Activer l'OCR par défaut
    
    # Paramètres pour le mode local
    DEFAULT_LOCAL_API_URL = "http://localhost:11434/v1/chat/completions"
    DEFAULT_LOCAL_MODEL = "deepseek-r1:1.5b"
    DEFAULT_LOCAL_MODE = False
    DEFAULT_CPU_ONLY = False
    
    # Ajoutez ces lignes après DEFAULT_CPU_ONLY
    DEFAULT_PROXY_ENABLED = False
    DEFAULT_PROXY_USERNAME = ""
    DEFAULT_PROXY_PASSWORD = ""
    DEFAULT_PROXY_HOST = ""
    DEFAULT_PROXY_PORT = ""

    # Paramètres pour Eden
    DEFAULT_EDEN_MODE = False
    DEFAULT_EDEN_API_URL = "https://edenportal-sfr.eden-genai.com/eden/ask"
    DEFAULT_EDEN_API_KEY = "eden-9b3e45fd08267f2b399f63e4e3f77ef7"
    DEFAULT_EDEN_GENI_ID = "b6742807-1efa-4416-8d8a-f72f3ad64e9b"
    DEFAULT_EDEN_INTERNET_VERIFY = False  # Option pour vérifier les informations sur internet

    # Listes de modèles courants
    COMMON_MODELS = [
        "google/gemini-2.0-flash-lite-preview-02-05:free",
        "google/gemini-2.0-flash-thinking-exp:free",
        "google/gemini-2.0-pro-exp-02-05:free",
        "deepseek/deepseek-r1:free",
        "deepseek/deepseek-r1-distill-llama-70b:free",
        "moonshotai/moonlight-16b-a3b-instruct:free",
        "nousresearch/deephermes-3-llama-3-8b-preview:free",
        "cognitivecomputations/dolphin3.0-r1-mistral-24b:free",
        "cognitivecomputations/dolphin3.0-mistral-24b:free",
        "mistralai/mistral-small-24b-instruct-2501:free",
        "qwen/qwen-vl-plus:free",
        "qwen/qwen2.5-vl-72b-instruct:free",
        "sophosympatheia/rogue-rose-103b-v0.2:free",
    ]

    # Pour mode local (Ollama)
    LOCAL_MODELS = [
        "deepseek-r1:1.5b",
        "llama3:8b",
        "llama3:70b",
        "mistral",
        "mixtral",
        "phi3:14b",
    ]

    THEMES = {
        "light": {
            "bg": "#f0f0f0",
            "fg": "black",
            "chat_bg": "#ffffff",
            "chat_fg": "black",
            "input_bg": "white",
            "input_fg": "black",
            "button_bg": "#d9d9d9",
            "button_fg": "black",
            "scroll_bg": "#f0f0f0",
            "scroll_fg": "black",
            "bubble_user_bg": "#DCF8C6",
            "bubble_user_fg": "black",
            "bubble_assistant_bg": "#f0f0f0",
            "bubble_assistant_fg": "black",
            "selectbackground": "#0078d7",
            "insertbackground": "black",
        },
        "dark": {
            "bg": "#2e2e2e",
            "fg": "white",
            "chat_bg": "#333333",
            "chat_fg": "white",
            "input_bg": "#444444",
            "input_fg": "white",
            "button_bg": "#555555",
            "button_fg": "white",
            "scroll_bg": "#444444",
            "scroll_fg": "white",
            "bubble_user_bg": "#4CAF50",
            "bubble_user_fg": "white",
            "bubble_assistant_bg": "#555555",
            "bubble_assistant_fg": "white",
            "selectbackground": "#0078d7",
            "insertbackground": "white",
        },
        "blue": {
            "bg": "#e6f2ff",
            "fg": "#003366",
            "chat_bg": "#ffffff",
            "chat_fg": "#003366",
            "input_bg": "white",
            "input_fg": "#003366",
            "button_bg": "#0066cc",
            "button_fg": "white",
            "scroll_bg": "#e6f2ff",
            "scroll_fg": "#003366",
            "bubble_user_bg": "#cce6ff",
            "bubble_user_fg": "#003366",
            "bubble_assistant_bg": "#e6e6ff",
            "bubble_assistant_fg": "#003366",
            "selectbackground": "#99ccff",
            "insertbackground": "#003366",
        },
    }

    @classmethod
    def load(cls):
        config = {
            "api_key": cls.DEFAULT_API_KEY,
            "api_url": cls.DEFAULT_API_URL,
            "model": cls.DEFAULT_MODEL,
            "typing_delay": cls.DEFAULT_TYPING_DELAY,
            "response_mode": cls.DEFAULT_RESPONSE_MODE,
            "window_width": cls.DEFAULT_WINDOW_WIDTH,
            "window_height": cls.DEFAULT_WINDOW_HEIGHT,
            "timeout": cls.DEFAULT_TIMEOUT,
            "theme": cls.DEFAULT_THEME,
            "local_mode": cls.DEFAULT_LOCAL_MODE,
            "cpu_only": cls.DEFAULT_CPU_ONLY,
            # Ajoutez ces lignes pour les paramètres de proxy
            "proxy_enabled": cls.DEFAULT_PROXY_ENABLED,
            "proxy_username": cls.DEFAULT_PROXY_USERNAME,
            "proxy_password": cls.DEFAULT_PROXY_PASSWORD,
            "proxy_host": cls.DEFAULT_PROXY_HOST,
            "proxy_port": cls.DEFAULT_PROXY_PORT,
            # Paramètres OCR
            "ocr_mode": cls.DEFAULT_OCR_MODE,
            "ocr_api_key": cls.DEFAULT_OCR_API_KEY,
            "ocr_lang": cls.DEFAULT_OCR_LANG,
            "ocr_enabled": cls.DEFAULT_OCR_ENABLED,
        }
        try:
            with open(cls.CONFIG_FILE, "r") as f:
                config.update(json.load(f))
                logging.info(f"Configuration chargée: {config}")
        except (FileNotFoundError, json.JSONDecodeError):
            logging.info(
                "Fichier de configuration non trouvé ou corrompu, utilisation des valeurs par défaut"
            )
        return config

    @classmethod
    def save(cls, config):
        try:
            with open(cls.CONFIG_FILE, "w") as f:
                json.dump(config, f, indent=4)
            logging.info("Configuration sauvegardée")
        except Exception as e:
            logging.error(f"Erreur lors de la sauvegarde de la configuration: {e}")
            messagebox.showerror(
                "Erreur", "Erreur lors de la sauvegarde de la configuration."
            )

    @classmethod
    def update_models(cls):
        """Met à jour la liste des modèles depuis OpenRouter"""
        try:
            response = requests.get(
                "https://openrouter.ai/api/v1/models",
                timeout=10
            )
            response.raise_for_status()
            models_data = response.json()
            
            # Filtrer les modèles gratuits
            free_models = []
            for model in models_data.get("data", []):
                model_id = model.get("id", "")
                if model_id.endswith(":free"):
                    free_models.append(model_id)
            
            if free_models:
                # Mettre à jour COMMON_MODELS avec les nouveaux modèles gratuits
                cls.COMMON_MODELS = free_models
                
                # Sauvegarder la nouvelle liste dans le fichier de configuration
                config = cls.load()
                config["available_models"] = free_models
                cls.save(config)
                
                return True, "Liste des modèles mise à jour avec succès"
            else:
                return False, "Aucun modèle gratuit trouvé"
                
        except requests.exceptions.RequestException as e:
            error_msg = f"Erreur lors de la mise à jour des modèles: {str(e)}"
            logging.error(error_msg)
            return False, error_msg
        except Exception as e:
            error_msg = f"Erreur inattendue: {str(e)}"
            logging.error(error_msg)
            return False, error_msg



def get_proxy_settings():
    """Récupère les paramètres de proxy à partir des variables d'environnement"""
    http_proxy = os.environ.get('HTTP_PROXY') or os.environ.get('http_proxy')
    https_proxy = os.environ.get('HTTPS_PROXY') or os.environ.get('https_proxy')
    no_proxy = os.environ.get('NO_PROXY') or os.environ.get('no_proxy')
    
    proxy_settings = {}
    if http_proxy:
        proxy_settings['http'] = http_proxy
        logging.info(f"Utilisation du proxy HTTP: {http_proxy}")
    if https_proxy:
        proxy_settings['https'] = https_proxy
        logging.info(f"Utilisation du proxy HTTPS: {https_proxy}")
    if no_proxy:
        proxy_settings['no_proxy'] = no_proxy
        logging.info(f"Exclusions proxy: {no_proxy}")



# ------------------- Vérification du serveur local -------------------


def is_port_open(url):
    try:
        parsed = urlparse(url)
        host = parsed.hostname
        port = parsed.port
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(5)
        s.connect((host, port))
        s.close()
        return True
    except Exception:
        return False


# ------------------- Classe pour gérer l'API -------------------


class AIClient:
    def __init__(self, api_key, api_url, model, timeout, proxy_config=None):
        self.api_key = api_key
        self.api_url = api_url
        self.model = model
        self.timeout = timeout
        self.current_request = None
        self.proxies = {}
        
        # Configurer le proxy selon les paramètres
        if proxy_config and proxy_config.get("proxy_enabled", False):
            username = proxy_config.get("proxy_username", "")
            password = proxy_config.get("proxy_password", "")
            host = proxy_config.get("proxy_host", "")
            port = proxy_config.get("proxy_port", "")
            
            if host and port:
                # Si username et password sont fournis, les inclure dans l'URL du proxy
                if username and password:
                    # Encoder correctement le mot de passe
                    encoded_password = urllib.parse.quote(password)
                    proxy_url = f"http://{username}:{encoded_password}@{host}:{port}"
                else:
                    proxy_url = f"http://{host}:{port}"
                
                self.proxies = {
                    'http': proxy_url,
                    'https': proxy_url
                }
                
                logging.info(f"Proxy configuré: {proxy_url}")
            else:
                logging.info("Proxy activé mais hôte ou port manquant")
        else:
            # Utiliser les paramètres de proxy système si disponibles
            self.proxies = get_proxy_settings()
            if self.proxies:
                logging.info("Paramètres de proxy système détectés et activés")
            else:
                logging.info("Aucun paramètre de proxy détecté")


    def chat_with_ai(self, user_message, message_history, model=None):
        # Log pour débogage
        logging.info(f"AIClient: Utilisation de {len(message_history)} messages dans l'historique")
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }

        messages = []
        for msg in message_history:
            role = msg.get("role", "")
            content = msg.get("content", "")
            if role and content:
                messages.append({"role": role, "content": content})

        # Ajouter le message actuel
        messages.append({"role": "user", "content": user_message})
        payload = {
            "model": model if model else self.model,
            "messages": messages,
        }

        # Log du contenu de la requête (pour déboguer)
        logging.info(f"Envoi à l'API avec {len(messages)} messages")

        try:
            self.current_request = requests.post(
                self.api_url,
                headers=headers,
                data=json.dumps(payload),
                timeout=self.timeout,
                proxies=self.proxies,
                verify=False  # Pour éviter les problèmes SSL avec les proxies
            )
            self.current_request.raise_for_status()
            response_data = self.current_request.json()
            self.current_request = None
            return response_data["choices"][0]["message"]["content"]
        except requests.exceptions.Timeout:
            logging.error("Timeout lors de l'appel à l'API")
            return "L'appel à l'API a expiré. Veuillez réessayer."
        except requests.exceptions.RequestException as e:
            logging.error(f"Erreur de connexion à l'API: {e}")
            return f"Erreur de connexion: {e}\nVérifiez votre clé API et votre URL."
        except (KeyError, json.JSONDecodeError) as e:
            logging.error(f"Erreur lors du traitement de la réponse de l'API: {e}")
            return "Erreur lors du traitement de la réponse de l'IA. Vérifiez votre clé API et votre URL."
        except Exception as e:
            logging.error(f"Une erreur inattendue s'est produite: {e}")
            return f"Une erreur s'est produite: {e}. Veuillez réessayer."

    def cancel_request(self):
        if self.current_request:
            try:
                self.current_request.close()
                logging.info("Requête annulée")
                return True
            except Exception as e:
                logging.error(f"Erreur lors de l'annulation de la requête: {e}")
        return False


# ------------------- Gestionnaire d'historique -------------------


class HistoryManager:
    def __init__(self, history_file):
        self.history_file = history_file
        self.message_history = self.load()

    def load(self):
        history = []
        try:
            with open(self.history_file, "r", encoding="utf-8") as f:
                history = json.load(f)
                logging.info(f"Historique chargé: {len(history)} messages")

                # Vérifier que l'historique a le bon format
                valid_history = []
                for msg in history:
                    if isinstance(msg, dict) and "role" in msg and "content" in msg:
                        valid_history.append(msg)
                    else:
                        logging.warning(f"Message ignoré car format invalide: {msg}")
            
                history = valid_history


        except (FileNotFoundError, json.JSONDecodeError):
            logging.info(
                "Historique non trouvé ou corrompu, création d'un nouvel historique"
            )
        return history

    def save(self):
        try:
            with open(self.history_file, "w", encoding="utf-8") as f:
                json.dump(self.message_history, f, indent=4)
                logging.info("Historique sauvegardé")
        except Exception as e:
            logging.error(f"Erreur lors de la sauvegarde de l'historique: {e}")
            messagebox.showerror(
                "Erreur", "Erreur lors de la sauvegarde de l'historique."
            )

    def add_message(self, role, content):
        logging.info(f"Ajout d'un message: {role} - {content[:20]}...")
        self.message_history.append({"role": role, "content": content})
        self.save()

    def clear(self):
        self.message_history = []
        self.save()


# ------------------- Classe de gestion de code -------------------


class CodeHandler:
    @staticmethod
    def extract_code(text):
        if not text:
            return None, None
        code_match = re.search(r"```(\w*)\n([\s\S]*?)\n```", text)
        if code_match:
            lang = code_match.group(1)
            code = code_match.group(2).strip()
            return lang, code
        return None, None

    @staticmethod
    def extract_excel_formula(text):
        """Extrait les formules Excel/VBA du texte"""
        if not text:
            return None
        
        # Recherche des formules Excel entre ```excel et ```
        excel_match = re.search(r"```excel\n([\s\S]*?)\n```", text)
        if excel_match:
            return "excel", excel_match.group(1).strip()
        
        # Recherche du code VBA entre ```vba et ```
        vba_match = re.search(r"```vba\n([\s\S]*?)\n```", text)
        if vba_match:
            return "vba", vba_match.group(1).strip()
        
        # Recherche des formules Excel qui commencent par =
        formula_match = re.search(r"\n=([\w\s\d\.\,\(\)\&\+\-\*\/\:\;\$\%\^\!\~\<\>\=]+)\n", text)
        if formula_match:
            return "formula", "=" + formula_match.group(1).strip()
        
        return None, None

    @staticmethod
    def highlight_code(code, language=None):
        if not code:
            return ""
        try:
            if (language and language.strip()):
                lexer = get_lexer_by_name(language)
            else:
                lexer = guess_lexer(code)
            formatter = get_formatter_by_name("html")
            highlighted_code = highlight(code, lexer, formatter)
            return highlighted_code
        except Exception as e:
            logging.error(f"Erreur de coloration syntaxique: {e}")
            return code

    @staticmethod
    def copy_to_clipboard(text):
        try:
            pyperclip.copy(text)
            messagebox.showinfo("Copie", "Texte copié dans le presse-papiers !")
            return True
        except Exception as e:
            logging.error(f"Erreur lors de la copie dans le presse-papiers: {e}")
            messagebox.showerror("Erreur", "Impossible de copier. Vérifiez pyperclip.")
            return False


# ------------------- Interface principale -------------------


class ChatApplication:
    def __init__(self, root):
        self.root = root
        self.config = Config.load()
        # Créer un dictionnaire de configuration du proxy
        proxy_config = {
            "proxy_enabled": self.config.get("proxy_enabled", False),
            "proxy_host": self.config.get("proxy_host", ""),
            "proxy_port": self.config.get("proxy_port", ""),
            "proxy_username": self.config.get("proxy_username", ""),
            "proxy_password": self.config.get("proxy_password", ""),
        }

        # Initialiser le client approprié selon le mode
        if self.config.get("eden_mode", False):
            self.ai_client = EdenClient(self.config["timeout"], proxy_config)
        else:
            self.ai_client = AIClient(
                self.config["api_key"],
                self.config["api_url"],
                self.config["model"],
                self.config["timeout"],
                proxy_config
            )

        self.history_manager = HistoryManager(Config.HISTORY_FILE)
        self.response_thread = None
        self.is_typing = False
        self.attachments = {}
        self.bold_text_enabled = False
        self.dual_chat_model = self.config.get("dual_chat_model", self.config["model"])

        self.setup_window()
        self.setup_ui()
        self.setup_menu()
        self.setup_status_bar()
        self.load_app_state()
        self.load_history_and_display()
        self.update_mode_indicator()

        # Initialiser le mode dual chat selon la configuration sauvegardée
        if self.config.get("dual_chat", False):
            self.create_dual_chat_box()
    
    def setup_window(self):
        self.root.title("Chat avec IA - Interface Améliorée par Marco")
        window_width = self.config.get("window_width", Config.DEFAULT_WINDOW_WIDTH)
        window_height = self.config.get("window_height", Config.DEFAULT_WINDOW_HEIGHT)
        self.root.geometry(f"{window_width}x{window_height}")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=0)
        self.root.rowconfigure(2, weight=0)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)


    def on_closing(self):
        if messagebox.askyesno("Quitter", "Voulez-vous quitter l'application ?"):
            if self.is_typing:
                self.is_typing = False
            self.ai_client.cancel_request()
            self.save_app_state()
            Config.save(self.config)
            self.history_manager.save()
            self.root.destroy()

    def setup_ui(self):
        """Configure l'interface utilisateur de l'application"""
        theme = Config.THEMES.get(self.config["theme"], Config.THEMES["light"])
        
        # Création du cadre principal pour le chat
        self.chat_frame = tk.Frame(self.root)
        self.chat_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        self.chat_frame.columnconfigure(0, weight=1)
        self.chat_frame.rowconfigure(0, weight=1)

        # Zone de texte principale pour l'affichage des messages
        self.chat_box = scrolledtext.ScrolledText(
            self.chat_frame,
            wrap=tk.WORD,
            font=("Arial", 10),
            borderwidth=2,
            relief="groove",
            state=tk.NORMAL,  # Toujours en mode NORMAL pour permettre la sélection
            cursor="xterm",  # Curseur visible
            takefocus=1,  # Peut recevoir le focus
        )
        self.chat_box.grid(row=0, column=0, sticky="nsew")

        # Configuration des couleurs de sélection native
        self.chat_box.config(
            selectbackground="#0000FF", selectforeground="white"  # Bleu vif
        )

        # Créer un tag personnalisé pour la surbrillance avec des couleurs très vives
        self.chat_box.tag_configure(
            "custom_highlight",
            background="#0000FF",  # Bleu vif
            foreground="white",
            borderwidth=1,
            relief="raised",
        )  # Effet légèrement en relief

        # Tag pour le texte en gras
        self.chat_box.tag_configure("bold", font=("Arial", 10, "bold"))

        # Définir les priorités des tags (custom_highlight a la plus haute priorité)
        self.chat_box.tag_raise("custom_highlight")
        self.chat_box.tag_raise("bold", "sel")  # Le gras est prioritaire sur la sélection standard

        # Liaisons d'événements étendues pour la sélection
        self.chat_box.bind("<<Selection>>", self.on_text_selected)
        self.chat_box.bind("<ButtonRelease-1>", self.on_text_selected)
        self.chat_box.bind("<KeyRelease>", self.on_text_selected)
        self.chat_box.bind("<FocusIn>", self.on_text_selected)
        self.chat_box.bind("<Control-a>", self.on_text_selected)  # Pour Ctrl+A (Tout sélectionner)

        # Autres liaisons pour garantir la sélection
        self.chat_box.bind(
            "<FocusIn>", lambda e: self.ensure_selectable(), add="+"
        )
        self.chat_box.bind("<Enter>", lambda e: self.ensure_selectable())

        # Scrollbar pour la zone de chat
        self.scrollbar = tk.Scrollbar(self.chat_frame, command=self.chat_box.yview)
        self.scrollbar.grid(row=0, column=1, sticky="ns")
        self.chat_box["yscrollcommand"] = self.scrollbar.set

        # Cadre pour la zone de saisie
        self.input_frame = tk.Frame(self.root)
        self.input_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        self.input_frame.columnconfigure(0, weight=1)

        # Zone de saisie utilisateur
        self.user_input = scrolledtext.ScrolledText(
            self.input_frame,
            height=4,
            font=("Arial", 10),
            borderwidth=2,
            relief="groove",
            insertbackground="black",
        )
        self.user_input.grid(row=0, column=0, sticky="ew")
        self.user_input.bind("<Button-1>", lambda event: self.user_input.focus_set())
        self.user_input.bind("<Return>", self.send_message)
        self.user_input.bind("<Control-Return>", self.insert_newline)

        # Zone pour afficher les fichiers joints
        self.attachments_frame = tk.Frame(self.input_frame)
        self.attachments_frame.grid(row=1, column=0, sticky="ew", pady=5)

        # Cadre pour les boutons d'action
        self.button_frame = tk.Frame(self.root)
        self.button_frame.grid(row=2, column=0, padx=10, pady=5, sticky="ew")

        # Configurer les colonnes pour les boutons (distribution égale)
        for i in range(4):  # 4 colonnes pour les 4 boutons
            self.button_frame.columnconfigure(i, weight=1)

        # Bouton Envoyer
        self.send_button = Button(
            self.button_frame, text="Envoyer", command=self.send_message, width=10 # <-- Ajout de width
        )
        self.send_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        # Bouton Annuler
        self.cancel_button = Button(
            self.button_frame,
            text="Annuler",
            command=self.cancel_generation,
            state=tk.NORMAL,
            width=10, # <-- Ajout de width
        )
        self.cancel_button.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        # Bouton pour joindre des fichiers
        self.attach_button = Button(
            self.button_frame, text="Joindre fichier", command=self.attach_file, width=15 # <-- Ajout de width
        )
        self.attach_button.grid(row=0, column=2, padx=5, pady=5, sticky="ew")


        # Étiquette pour l'indicateur de chargement
        self.loading_label = Label(self.root, text="", font=("Arial", 9, "italic"))

        # Configuration du menu contextuel et du thème
        self.setup_context_menu()
        self.apply_theme(self.config["theme"])

        # Forcer la mise à jour pour que les styles soient appliqués
        self.root.update_idletasks()
        self.chat_box.bind("<Button-3>", self.show_context_menu)  # <<-- C'est ICI que vous créez le menu
        self.setup_context_menu()
        
        # Ajout de la case à cocher pour le mode Dual chat_box
        self.dual_chat_var = tk.BooleanVar(value=self.config.get("dual_chat", Config.DEFAULT_DUAL_CHAT))
        self.dual_chat_checkbox = tk.Checkbutton(
            self.button_frame,
            text="Mode Dual chat_box",
            variable=self.dual_chat_var,
            command=self.toggle_dual_chat,
            bg=theme["bg"],
            fg=theme["fg"],
            selectcolor=theme["bg"]
        )
        self.dual_chat_checkbox.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
        
        
    def process_excel_request(self, assistant_response, attached_file_path=None):
        """
        Traite une demande de création/modification de fichier Excel basée sur 
        la réponse de l'assistant et un fichier joint si disponible.
        """
        if pd is None:
            messagebox.showerror(
                "Erreur",
                "La bibliothèque pandas est requise pour le traitement Excel. Veuillez l'installer avec 'pip install pandas'."
                )
            return False
    
        try:
            # Vérifier si la réponse contient du code Python pour manipuler des données
            code_pattern = r"```python\s*([\s\S]*?)\s*```"
            match = re.search(code_pattern, assistant_response)
        
            if not match:
                logging.warning("Aucun code Python trouvé dans la réponse de l'assistant")
                return False
        
            code = match.group(1)
        
            # Si un fichier Excel a été joint, le rendre disponible au code
            input_data = None
            if attached_file_path and os.path.exists(attached_file_path):
                try:
                    if attached_file_path.endswith('.csv'):
                        input_data = pd.read_csv(attached_file_path)
                    else:
                            input_data = pd.read_excel(attached_file_path)
                            logging.info(f"Fichier Excel/CSV lu avec succès: {attached_file_path}")
                except Exception as e:
                    logging.error(f"Erreur lors de la lecture du fichier joint: {e}")
                    messagebox.showerror("Erreur", f"Impossible de lire le fichier Excel: {e}")
                    return False
        
            # Préparer l'environnement d'exécution sécurisé
            exec_globals = {
                'pd': pd,
                'input_data': input_data,
                'result_data': None,
                'print': print  # Pour le débogage
                }
        
            # Vérifier que le code ne contient pas d'opérations dangereuses
            dangerous_patterns = ['os.', 'subprocess.', 'sys.', 'eval(', 'exec(', '__import__']
            for pattern in dangerous_patterns:
                if pattern in code:
                    messagebox.showerror(
                        "Erreur de sécurité", 
                        f"Le code contient des opérations potentiellement dangereuses: {pattern}"
                        )
                    return False
        
            # Modifier le code pour assurer qu'il assigne le résultat à 'result_data'
            if 'result_data' not in code:
                # Ajouter une ligne à la fin pour stocker le résultat dans result_data
                # Rechercher le dernier DataFrame créé
                df_pattern = r"(\w+)\s*=\s*(?:pd\.DataFrame|input_data|.*DataFrame)"
                df_matches = re.finditer(df_pattern, code)
                last_df = None
                for match in df_matches:
                    last_df = match.group(1)
                    
                    if last_df:
                        code += f"\n# Assigner automatiquement le dernier DataFrame comme résultat\nresult_data = {last_df}"
                    else:
                        messagebox.showerror(
                            "Erreur", 
                            "Impossible de détecter un DataFrame dans le code. Le code doit créer un DataFrame comme résultat."
                            )
                        return False
        
            # Exécuter le code
            try:
                exec(code, exec_globals)
                result_data = exec_globals.get('result_data')
            
                if result_data is None or not isinstance(result_data, pd.DataFrame):
                    messagebox.showerror(
                        "Erreur",
                        "Le code n'a pas produit de DataFrame valide. Assurez-vous que le code assigne un DataFrame à 'result_data'."
                        )
                    return False
            
                # Demander où sauvegarder le fichier résultat
                output_file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[
                        ("Fichier Excel", "*.xlsx"),
                        ("Fichier CSV", "*.csv"),
                        ],
                    title="Enregistrer le fichier traité"
                    )
            
                if not output_file_path:
                    return False  # L'utilisateur a annulé
            
                # Sauvegarder le résultat
                if output_file_path.endswith('.csv'):
                    result_data.to_csv(output_file_path, index=False)
                else:
                        result_data.to_excel(output_file_path, index=False)
            
                messagebox.showinfo(
                    "Traitement terminé",
                    f"Le fichier a été traité avec succès et enregistré sous:\n{output_file_path}"
                    )
                return True
            
            except Exception as e:
                logging.error(f"Erreur lors de l'exécution du code: {e}")
                messagebox.showerror("Erreur d'exécution", f"Une erreur s'est produite lors du traitement: {e}")
                return False
    
        except Exception as e:
            logging.error(f"Erreur lors du traitement de la demande Excel: {e}")
            messagebox.showerror("Erreur", f"Une erreur inattendue s'est produite: {e}")
            return False    
    
    
    def detect_excel_processing_request(self, response_text):
        """
        Détecte si la réponse contient du code Python pour traiter un fichier Excel
        et propose à l'utilisateur d'exécuter ce traitement.
        """
        # Vérifier si la réponse contient du code Python pour pandas
        if "```python" in response_text and ("pd." in response_text or "pandas" in response_text):
            # Vérifier si le code semble traiter des fichiers Excel
            excel_indicators = [
                "read_excel", "to_excel", "read_csv", "to_csv", 
                "DataFrame", "drop_duplicates", "fillna", "groupby",
                "pivot_table", "sort_values", "filter"
                ]
        
            for indicator in excel_indicators:
                if indicator in response_text:
                    # Vérifier si un fichier Excel a été joint récemment
                    attached_file_path = None
                    for attachment_id, attachment in self.attachments.items():
                        if attachment["path"].lower().endswith(('.xlsx', '.xls', '.csv')):
                            attached_file_path = attachment["path"]
                            break
                
                # Si aucun fichier n'est actuellement joint mais qu'il y a du code de traitement,
                # chercher dans l'historique récent
                if not attached_file_path:
                    # On pourrait ajouter un mécanisme pour retrouver le dernier fichier
                    # Excel mentionné dans la conversation
                    pass
                
                # Proposer à l'utilisateur d'exécuter le traitement
                if messagebox.askyesno(
                    "Traitement Excel détecté",
                    "L'assistant a généré du code pour traiter un fichier Excel. Voulez-vous exécuter ce traitement maintenant ?"
                ):
                    self.process_excel_request(response_text, attached_file_path)
                
                return True
        
            return False
    
        return False                
    
    
    
    
    def toggle_dual_chat(self):
        is_dual = self.dual_chat_var.get()
        self.config["dual_chat"] = is_dual
        
        if is_dual:
            # Créer un client AI temporaire pour le dual chat avec le modèle par défaut
            self.dual_chat_model = self.config.get("dual_chat_model", Config.DEFAULT_MODEL)
            self.create_dual_chat_box()
        else:
            # Désactiver le dual chat et supprimer le deuxième modèle
            self.dual_chat_model = None
            self.remove_dual_chat_box()
        
        # Sauvegarder la configuration
        Config.save(self.config)
        
        # Mettre à jour l'indicateur de mode
        self.update_mode_indicator()
    
    def create_dual_chat_box(self):
        # Récupérer le thème courant
        theme = Config.THEMES.get(self.config["theme"], Config.THEMES["light"])

        # Créer une deuxième zone de chat
        self.chat_box2 = scrolledtext.ScrolledText(
            self.chat_frame,
            wrap=tk.WORD,
            font=("Arial", 10),
            borderwidth=2,
            relief="groove",
            state=tk.NORMAL,
            cursor="xterm",
            takefocus=1
        )
        self.chat_box2.grid(row=0, column=1, sticky="nsew")
        self.chat_frame.columnconfigure(1, weight=1)

        # Appliquer le thème à la deuxième fenêtre
        self.chat_box2.config(
            bg=theme["chat_bg"],
            fg=theme["chat_fg"],
            insertbackground=theme["insertbackground"],
            selectbackground="#0000FF",  # Bleu vif pour la sélection
            selectforeground="white"
        )

        # Configurer les tags pour la deuxième fenêtre
        self.chat_box2.tag_configure(
            "user",
            foreground=theme["bubble_user_fg"],
            background=theme["bubble_user_bg"]
        )
        self.chat_box2.tag_configure(
            "assistant",
            foreground=theme["bubble_assistant_fg"],
            background=theme["bubble_assistant_bg"]
        )

        # Configurer le tag de surbrillance pour la deuxième fenêtre
        self.chat_box2.tag_configure(
            "custom_highlight",
            background="#0000FF",  # Bleu vif
            foreground="white",
            borderwidth=1,
            relief="raised"
        )

        # Tag pour le texte en gras dans la deuxième fenêtre
        self.chat_box2.tag_configure("bold", font=("Arial", 10, "bold"))
    
        # Ajouter les mêmes bindings pour la sélection
        self.chat_box2.bind("<<Selection>>", self.on_text_selected)
        self.chat_box2.bind("<ButtonRelease-1>", self.on_text_selected)
        self.chat_box2.bind("<KeyRelease>", self.on_text_selected)
        self.chat_box2.bind("<FocusIn>", self.on_text_selected)
        self.chat_box2.bind("<Control-a>", self.on_text_selected)


    def remove_dual_chat_box(self):
        if hasattr(self, 'chat_box2'):
            self.chat_box2.destroy()
            delattr(self, 'chat_box2')
        self.chat_frame.columnconfigure(1, weight=0)
    
   
    
    def on_text_selected(self, event=None):
        """Méthode appelée automatiquement lors de la sélection de texte"""
        # Déboguer - Vous pouvez décommenter cette ligne pour vérifier que l'événement est bien capturé
        # print("Sélection détectée!")

        # Attendre un court instant pour que la sélection soit complètement établie
        self.root.after(10, self._apply_selection_style)
        return None  # Ne pas interrompre la propagation des événements   
        
        

    def extract_excel_preview(self, file_path, max_rows_to_read=10, max_cols_to_read=10):
        """Extrait un aperçu du contenu d'un fichier Excel avec limite explicite des lignes et colonnes"""
        try:
            # Vérifier si openpyxl est disponible
            import openpyxl
            
            # Enregistrer un message de débogage
            logging.info(f"Ouverture du fichier Excel: {file_path}")
            
            # Ouvrir le classeur Excel
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
            # Récupérer les noms des feuilles
            sheet_names = workbook.sheetnames
            logging.info(f"Feuilles trouvées: {sheet_names}")
        
            preview = {
                "sheets": sheet_names,
                "first_rows": {}
            }
            
            # Pour chaque feuille, récupérer les données
            for sheet_name in sheet_names:
                logging.info(f"Analyse de la feuille: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # Estimation simplifiée du nombre de lignes
                try:
                    # Évaluation rapide de la taille
                    if hasattr(sheet, 'max_row') and sheet.max_row > 1000:
                        total_rows = "plus de 1000 lignes"
                    else:
                        row_count = 0
                        for _ in sheet.iter_rows(min_row=1, max_row=1000):
                            row_count += 1
                        
                        if row_count == 1000:  # Si on a atteint notre limite de test
                            total_rows = "plus de 1000 lignes"
                        else:
                            total_rows = f"{row_count} lignes"
                except Exception as e:
                    logging.error(f"Erreur lors du comptage des lignes: {e}")
                    total_rows = "nombre inconnu"
                
                logging.info(f"Nombre total de lignes estimé: {total_rows}")
                
                # Récupérer les en-têtes (limités à max_cols_to_read)
                headers = []
                try:
                    header_row = None
                    for row in sheet.iter_rows(min_row=1, max_row=1, max_col=max_cols_to_read, values_only=True):
                        header_row = row
                        break
                    
                    if header_row:
                        headers = [str(cell) if cell is not None else "" for cell in header_row]
                        logging.info(f"En-têtes trouvés: {headers}")
                    else:
                        logging.warning("Aucun en-tête trouvé dans la première ligne")
                        headers = [f"Colonne {i+1}" for i in range(max_cols_to_read)]
                except Exception as e:
                    logging.error(f"Erreur lors de la lecture des en-têtes: {e}")
                    headers = [f"Colonne {i+1}" for i in range(max_cols_to_read)]
                
                # Récupérer les premières lignes (limitées)
                data_rows = []
                try:
                    rows_read = 0
                    # Commencer à partir de la ligne 2 (après les en-têtes)
                    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=max_rows_to_read+1, max_col=max_cols_to_read, values_only=True)):
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        data_rows.append(row_data)
                        rows_read += 1
                        logging.info(f"Lecture ligne {i+2}: {row_data}")
                        if rows_read >= max_rows_to_read:
                            break
                    logging.info(f"Nombre de lignes de données lues: {len(data_rows)}")
                except Exception as e:
                    logging.error(f"Erreur lors de la lecture des données: {e}", exc_info=True)
                
                # Indiquer si les données ont été tronquées
                has_more_cols = False
                if hasattr(sheet, 'max_column') and sheet.max_column > max_cols_to_read:
                    has_more_cols = True
                
                preview["first_rows"][sheet_name] = {
                    "headers": headers,
                    "sample_rows": data_rows,
                    "total_rows": total_rows,
                    "has_more_columns": has_more_cols,
                    "columns_shown": min(max_cols_to_read, len(headers))
                }
            
            workbook.close()
            logging.info(f"Aperçu Excel généré avec succès: {preview}")
            return preview

        except ImportError:
            logging.error("Le module openpyxl n'est pas disponible. Installez-le avec 'pip install openpyxl'")
            return {"sheets": [], "first_rows": {}, "error": "Le module openpyxl n'est pas disponible."}
        except Exception as e:
            logging.error(f"Erreur lors de l'extraction de l'aperçu Excel: {e}", exc_info=True)
            return {"sheets": [], "first_rows": {}, "error": str(e)}
 
        
    def send_message(self, event=None):
  
        """Envoie un message avec les pièces jointes éventuelles"""
        user_message = self.user_input.get("1.0", END).strip()
        if not user_message and not self.attachments:
            return "break" if event else None
        
         # Ajouter cette condition pour gérer correctement l'événement Entrée
        if event and event.keysym == 'Return' and not event.state & 4:  # Sans modifier Ctrl
            if not user_message and not self.attachments:
                return "break"  # Empêcher l'insertion d'une nouvelle ligne si vide
    
        if not user_message and not self.attachments:
            return "break" if event else None


        logging.info(f"État de l'historique avant envoi : {len(self.history_manager.message_history)} messages")
        for i, msg in enumerate(self.history_manager.message_history[-5:] if len(self.history_manager.message_history) > 5 else self.history_manager.message_history):
            logging.info(f"Message {i}: {msg.get('role')} - {msg.get('content')[:50]}...")

        # Créer un dictionnaire de configuration du proxy
        proxy_config = {
            "proxy_enabled": self.config.get("proxy_enabled", False),
            "proxy_host": self.config.get("proxy_host", ""),
            "proxy_port": self.config.get("proxy_port", ""),
            "proxy_username": self.config.get("proxy_username", ""),
            "proxy_password": self.config.get("proxy_password", ""),
            }

        # Préparer un message avec les pièces jointes si elles existent
        full_message = user_message

        if self.attachments:
            full_message = user_message + "\n\n[Pièces jointes]\n"
            for file_id, attachment in self.attachments.items():
                file_name = attachment['name']
                file_type = attachment['type']
                file_path = attachment["path"]

                if file_path.endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif')):
                    if self.config.get("ocr_enabled", True):
                        ocr_text = OCRHandler.extract_text(file_path, 
                                                          lang=self.config.get("ocr_lang", "fra+eng"),
                                                          use_online=self.config.get("ocr_mode", "online") == "online",
                                                          proxy_config=proxy_config
                                                          )
                        if ocr_text and ocr_text != "Aucun texte n'a pu être extrait de l'image.":
                            full_message += f"\nTexte extrait de l'image {file_name}:\n```\n{ocr_text}\n```\n"

                # Traitement spécial pour les fichiers Python
                elif file_path.endswith('.py') or file_type == "application/python":
                    try:
                        with open(file_path, "r", encoding="utf-8") as f:
                            file_content = f.read()
                            full_message += f"\nContenu du fichier Python {file_name}:\n```python\n{file_content}\n```\n"
                            full_message += "\nVeuillez analyser ce code Python et répondre à mes questions à son sujet.\n"
                    except Exception as e:
                        logging.error(f"Erreur lors de la lecture du fichier Python {file_path}: {e}")
                        full_message += f"\nErreur lors de la lecture de {file_name}: {str(e)}\n"
                
                # Traitement spécial pour les fichiers Excel
                elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                    try:
                        # Extraire l'aperçu Excel
                        excel_preview = self.extract_excel_preview(file_path, max_rows_to_read=10, max_cols_to_read=15)
                        
                        if "error" in excel_preview:
                            full_message += f"- {file_name} ({file_type}) - Erreur: {excel_preview['error']}\n"
                        else:
                            # Ajouter des informations détaillées sur le contenu Excel
                            sheets = excel_preview.get("sheets", [])
                            full_message += f"- {file_name} ({file_type}) - {len(sheets)} feuille(s)\n"
                            
                            # Pour chaque feuille, ajouter un aperçu des données
                            for sheet_name in sheets:
                                if sheet_name in excel_preview["first_rows"]:
                                    sheet_data = excel_preview["first_rows"][sheet_name]
                                    headers = sheet_data.get("headers", [])
                                    sample_rows = sheet_data.get("sample_rows", [])
                                    total_rows = sheet_data.get("total_rows", "inconnu")
                                    
                                    full_message += f"\nFeuille '{sheet_name}' ({total_rows}):\n"
                                    
                                    # Afficher les en-têtes
                                    if headers:
                                        full_message += "| " + " | ".join(headers) + " |\n"
                                        full_message += "| " + " | ".join(["---" for _ in headers]) + " |\n"
                                        
                                        # Afficher les premières lignes
                                        for row in sample_rows:
                                            # S'assurer que la ligne a la même longueur que les en-têtes
                                            padded_row = row + [""] * (len(headers) - len(row))
                                            full_message += "| " + " | ".join(padded_row) + " |\n"
                                        
                                        if sheet_data.get("has_more_columns", False):
                                            full_message += "(d'autres colonnes sont disponibles...)\n"
                            
                            full_message += "\nVeuillez analyser ces données Excel et répondre à mes questions à leur sujet.\n"
                    except Exception as e:
                        logging.error(f"Erreur lors de la lecture du fichier Excel {file_path}: {e}", exc_info=True)
                        full_message += f"\nErreur lors de la lecture de {file_name}: {str(e)}\n"
                else:
                    # Pour les autres types de fichiers
                    full_message += f"- {file_name} ({file_type})\n"

                    if file_type and (
                            file_type.startswith("text/") or
                            file_type == "text/csv" or
                            file_path.endswith(".txt") or
                            file_path.endswith(".csv")
                    ):
                        try:
                            with open(file_path, "r", encoding="utf-8") as f:
                                file_content = f.read(5000000)  # Limite de 5MB
                                if f.read(1):  # S'il reste du contenu
                                    file_content += "\n... (contenu tronqué)"
                                full_message += f"\nContenu de {file_name}:\n```\n{file_content}\n```\n"
                        except Exception as e:
                            logging.error(f"Erreur lors de la lecture du fichier {file_path}: {e}")
                            full_message += f"\nErreur lors de la lecture de {file_name}: {str(e)}\n"

        logging.info(f"Message envoyé avec pièces jointes: {full_message[:200]}...")
        self.user_input.delete("1.0", END)
        self.append_message("user", full_message)
        self.history_manager.add_message("user", full_message)

        # Effacer les pièces jointes après l'envoi
        for widget in self.attachments_frame.winfo_children():
            widget.destroy()
        self.attachments.clear()

        # Lancer la génération de réponse
        self.start_loading()
        self.response_thread = threading.Thread(
            target=self.fetch_and_display_response, args=(full_message,)
        )
        self.response_thread.daemon = True
        self.response_thread.start()

        return "break" if event else None

    def _apply_selection_style(self):
        """Applique les styles à la sélection actuelle avec des couleurs forcées"""
        try:
            # Vérifier si une sélection existe
            if self.chat_box.tag_ranges("sel"):
                # Garantir que le mode sélection fonctionne
                self.ensure_selectable()

                # Récupérer les indices de début et de fin de la sélection
                start = self.chat_box.index("sel.first")
                end = self.chat_box.index("sel.last")

                # Nettoyer d'abord les anciennes surbrillances personnalisées
                self.chat_box.tag_remove("custom_highlight", "1.0", "end")

                # Appliquer notre propre surbrillance avec priorité maximale
                self.chat_box.tag_add("custom_highlight", start, end)

                # Appliquer le mode gras si activé
                if self.bold_text_enabled:
                    # Appliquer le gras sur la sélection
                    self.chat_box.tag_add("bold", start, end)

                # Forcer la mise à jour de la sélection et des tags
                self.chat_box.tag_raise("custom_highlight")  # Réappliquer la priorité

                # Forcer la mise à jour de l'affichage
                self.root.update_idletasks()

                # Déboguer - Vous pouvez décommenter ces lignes pour vérifier les indices
                # print(f"Sélection de {start} à {end}")
                # print(f"Tags actifs: {self.chat_box.tag_names()}")

        except tk.TclError as e:
            logging.error(f"Erreur lors de la sélection de texte: {e}")
        except Exception as e:
            logging.error(f"Erreur inattendue lors de la sélection: {e}")

    def ensure_selectable(self):
        """S'assure que le widget chat_box est sélectionnable"""
        self.chat_box.config(state=tk.NORMAL)
        return True

    def force_highlight(self):
        """Force la surbrillance de la sélection actuelle"""
        try:
            self.chat_box.update_idletasks()  # Mettre à jour l'état du widget
            self._apply_selection_style()  # Appliquer les styles de sélection
        except Exception as e:
            logging.error(f"Erreur lors du forçage de la surbrillance: {e}")

    def toggle_bold_mode(self):
        """Active/désactive le mode gras pour tout le texte."""
        self.bold_text_enabled = not self.bold_text_enabled
        theme = Config.THEMES.get(self.config["theme"], Config.THEMES["light"])

        if self.bold_text_enabled:
           self.chat_box.tag_add("bold", "1.0", END)  # Appliquer à tout le texte

        else:
            self.chat_box.tag_remove("bold", "1.0", END)  # Retirer de tout le texte


    def apply_bold_to_existing_messages(self):
        """Applique le mode gras aux messages existants"""
        if self.bold_text_enabled:
            # Parcourir chaque ligne du chat_box
            for i in range(1, float(self.chat_box.index(END))): # Commencer à 1.0 pour ignorer la ligne vide initiale
                try:
                    # Obtenir la plage de texte de chaque ligne
                    start = f"{i}.0"
                    end = f"{i}.end"

                    # Vérifier si la ligne a du texte
                    if self.chat_box.get(start, end).strip():
                        self.chat_box.tag_add("bold", start, end) # Appliquer tag
                except tk.TclError:
                    pass # Ignorer les erreurs liées à l'index

    def remove_bold_from_existing_messages(self):
        """Enlève le mode gras aux messages existants"""
        # Supprimer le gras de tout le texte
        self.chat_box.tag_remove("bold", "1.0", "end")
        self.force_highlight() # Refresh
        self.root.update_idletasks()
            # Mais garder la surbrillance
        #self.force_highlight()

    def update_selection_colors(self):
        self.chat_box.tag_configure("sel", background="blue", foreground="white")
        # Mettre à jour l'apparence
        self.root.update_idletasks()

    def setup_status_bar(self):
        self.status_label = Label(
            self.root, text="", anchor="w", bd=1, relief="sunken"
        )
        self.status_label.grid(row=3, column=0, sticky="ew")

    def update_mode_indicator(self):
        if self.config.get("eden_mode", False):
            mode = "Eden AI"
        elif self.config.get("local_mode", False):
            mode = "Local"
            if self.config.get("cpu_only"):
                mode += " (CPU only)"
        else:
            mode = "Online"
            mode += f" (Modèle: {self.config['model']})"
        
        # Ajouter l'indicateur de proxy
        if self.config.get("proxy_enabled", False):
            mode += " | Proxy: Activé"
        
        # Ajouter l'indicateur de dual chat si activé
        if self.config.get("dual_chat", False):
            mode += f" | Dual: {self.dual_chat_model}"

        self.status_label.config(text=f"Mode: {mode}")

    def setup_menu(self):
        self.menubar = Menu(self.root)
        self.filemenu = Menu(self.menubar, tearoff=0)
        self.filemenu.add_command(
            label="Nouvelle conversation", command=self.new_conversation
        )
        self.filemenu.add_command(
            label="Charger l'historique", command=self.load_history_and_display
        )
        self.filemenu.add_command(
            label="Sauvegarder l'historique",
            command=lambda: self.history_manager.save(),
        )
        self.filemenu.add_command(
            label="Effacer l'historique", command=self.clear_history
        )
        self.filemenu.add_separator()
        self.filemenu.add_command(
            label="Configuration", command=self.open_config_window
        )
        self.filemenu.add_command(
            label="Passer en mode local", command=self.switch_to_local_mode
        )
        self.filemenu.add_command(
            label="Passer en mode distant", command=self.switch_to_remote_mode
        )
        self.filemenu.add_command(
            label="Passer en mode Eden", command=self.switch_to_eden_mode
        )
        self.filemenu.add_separator()
        self.filemenu.add_command(label="Quitter", command=self.on_closing)
        self.menubar.add_cascade(label="Fichier", menu=self.filemenu)

        self.editmenu = Menu(self.menubar, tearoff=0)
        self.editmenu.add_command(label="Copier", command=self.copy_selected_text)
        self.editmenu.add_command(
            label="Coller", command=lambda: self.user_input.event_generate("<<Paste>>")
        )
        self.editmenu.add_separator()
        self.editmenu.add_command(label="Rechercher", command=self.search_text)
        self.menubar.add_cascade(label="Edition", menu=self.editmenu)

        self.viewmenu = Menu(self.menubar, tearoff=0)
        self.thememenu = Menu(self.viewmenu, tearoff=0)
        for theme in Config.THEMES:
            self.thememenu.add_command(
                label=theme.capitalize(), command=lambda t=theme: self.change_theme(t)
            )
        self.viewmenu.add_cascade(label="Thème", menu=self.thememenu)
        self.viewmenu.add_separator()
        self.viewmenu.add_command(label="Mode Gras", command=self.toggle_bold_mode)
        self.font_size_var = tk.IntVar(value=10)
        self.viewmenu.add_command(
            label="Augmenter la taille de police", command=self.increase_font_size
        )
        self.viewmenu.add_command(
            label="Diminuer la taille de police", command=self.decrease_font_size
        )
        self.menubar.add_cascade(label="Affichage", menu=self.viewmenu)

        self.helpmenu = Menu(self.menubar, tearoff=0)
        self.helpmenu.add_command(label="À propos", command=self.show_about)
        self.helpmenu.add_command(label="Aide", command=self.show_help)
        self.menubar.add_cascade(label="Aide", menu=self.helpmenu)

        self.root.config(menu=self.menubar)
        self.root.bind("<Control-f>", lambda event: self.search_text())
        self.root.bind("<Control-n>", lambda event: self.new_conversation())
        self.root.bind("<Control-s>", lambda event: self.history_manager.save())

    def setup_context_menu(self):
        self.chat_menu = Menu(self.root, tearoff=0)
        self.chat_menu.add_command(label="Copier", command=self.copy_selected_text)
        
        self.chat_menu.add_command(
            label="Copier code", command=self.copy_code_from_chat
        )
        self.chat_menu.add_command(
            label="Tout sélectionner",
            command=lambda: self.chat_box.tag_add("sel", "1.0", "end"),
        )
        self.chat_menu.add_command(label="Copier formule Excel/VBA", command=self.copy_excel_from_chat)
        self.chat_menu.add_command(label="Copier l'email", command=self.detect_and_copy_email)
        self.chat_box.bind("<Button-3>", self.show_context_menu)

    def show_context_menu(self, event):
        try:
            self.chat_menu.tk_popup(event.x_root, event.y_root, 0)
        finally:
            self.chat_menu.grab_release()

    def apply_theme(self, theme_name):
        """Applique le thème sélectionné à l'interface"""
        theme = Config.THEMES.get(theme_name, Config.THEMES["light"])

        # Appliquer les couleurs de base à l'interface
        self.root.config(bg=theme["bg"])
        self.chat_frame.config(bg=theme["bg"])
        self.input_frame.config(bg=theme["bg"])
        self.button_frame.config(bg=theme["bg"])
        self.attachments_frame.config(bg=theme["bg"])

        # Appliquer les couleurs au chat_box
        self.chat_box.config(
            bg=theme["chat_bg"],
            fg=theme["chat_fg"],
            insertbackground=theme["insertbackground"],
            selectbackground="#0000FF",  # Forcer le bleu vif pour la sélection native
            selectforeground="white",  # Forcer le blanc pour le texte sélectionné
        )

        # Appliquer le thème à la deuxième fenêtre de chat si elle existe
        if hasattr(self, 'chat_box2'):
            self.chat_box2.config(
                bg=theme["chat_bg"],
                fg=theme["chat_fg"],
                insertbackground=theme["insertbackground"],
                selectbackground="#0000FF",  # Bleu vif pour la sélection
                selectforeground="white"
            )
        
            # Mettre à jour les tags pour la deuxième fenêtre
            self.chat_box2.tag_config(
                "user",
                foreground=theme["bubble_user_fg"],
                background=theme["bubble_user_bg"]
            )
            self.chat_box2.tag_config(
                "assistant",
                foreground=theme["bubble_assistant_fg"],
                background=theme["bubble_assistant_bg"]
            )
        
            # Maintenir la même configuration pour la surbrillance
            self.chat_box2.tag_configure(
                "custom_highlight",
                background="#0000FF",
                foreground="white",
                borderwidth=1,
                relief="raised"
            )

        # Appliquer les couleurs à la zone de saisie
        self.user_input.config(
            bg=theme["input_bg"],
            fg=theme["input_fg"],
            insertbackground=theme["insertbackground"],
        )

        # Appliquer les couleurs aux boutons
        self.send_button.config(bg=theme["button_bg"], fg=theme["button_fg"])
        self.cancel_button.config(bg=theme["button_bg"], fg=theme["button_fg"])
        self.attach_button.config(bg=theme["button_bg"], fg=theme["button_fg"])

        # Appliquer les couleurs à la scrollbar
        self.scrollbar.config(bg=theme["scroll_bg"], troughcolor=theme["input_bg"])

        # Appliquer les couleurs à l'étiquette de chargement
        self.loading_label.config(bg=theme["bg"], fg=theme["fg"])

        # Appliquer les couleurs aux bulles de message
        self.chat_box.tag_config(
            "user",
            foreground=theme["bubble_user_fg"],
            background=theme["bubble_user_bg"],
        )
        self.chat_box.tag_config(
            "assistant",
            foreground=theme["bubble_assistant_fg"],
            background=theme["bubble_assistant_bg"],
        )

        # Maintenir les mêmes configurations pour la surbrillance personnalisée
        # indépendamment du thème pour garantir la visibilité
        self.chat_box.tag_configure(
            "custom_highlight",
            background="#0000FF",  # Bleu vif
            foreground="white",
            borderwidth=1,
            relief="raised",
        )

        # Rétablir les priorités des tags
        self.chat_box.tag_raise("custom_highlight")
        self.chat_box.tag_raise("bold", "sel")

        # Mettre à jour le thème dans la configuration
        self.config["theme"] = theme_name

        # Forcer la mise à jour de l'interface
        self.root.update_idletasks()

        # Si du texte est actuellement sélectionné, réappliquer la surbrillance
        self.force_highlight()

    def increase_font_size(self):
        current_size = self.font_size_var.get()
        if current_size < 24:
            new_size = current_size + 1
            self.font_size_var.set(new_size)
            self.chat_box.config(font=("Arial", new_size))
            
    def auto_highlight_selection(self, event=None):
        try:
            start = self.chat_box.index("sel.first")
            end = self.chat_box.index("sel.last")

            # Appliquer un tag temporaire
            self.chat_box.tag_add("highlight", start, end)
            self.chat_box.tag_config("highlight", background="yellow", foreground="black")

            # Programmer la suppression de la surbrillance après un délai (3 secondes)
            self.root.after(3000, lambda: self.chat_box.tag_remove("highlight", "1.0", END))

        except tk.TclError:
            # Pas de sélection, on ne fait rien.
            pass
        except Exception as e:
            logging.error(f"Erreur lors de la surbr	illance automatique: {e}")
    def decrease_font_size(self):
        current_size = self.font_size_var.get()
        if current_size > 8:
            new_size = current_size - 1
            self.font_size_var.set(new_size)
            self.chat_box.config(font=("Arial", new_size))
            self.user_input.config(font=("Arial", new_size))

    def change_theme(self, theme_name):
        self.apply_theme(theme_name)
        Config.save(self.config)

    def new_conversation(self):
        if self.history_manager.message_history:
            if messagebox.askyesno(
                "Nouvelle conversation",
                "Voulez-vous commencer une nouvelle conversation ? L'historique actuel sera sauvegardé.",
            ):
                self.history_manager.save()
                self.history_manager.clear()
                self.clear_chat_display()

    def clear_chat_display(self):
        self.chat_box.delete("1.0", END)
        if hasattr(self, 'chat_box2'):
            self.chat_box2.delete("1.0", END)
        
    def clear_history(self):
        if messagebox.askyesno(
            "Effacer l'historique",
            "Êtes-vous sûr de vouloir effacer tout l'historique de conversation ?",
        ):
            self.history_manager.clear()
            self.clear_chat_display()
            messagebox.showinfo("Historique", "Historique effacé.")

    def load_history_and_display(self):
        """Charge l'historique des messages et les affiche dans les fenêtres appropriées"""
        self.clear_chat_display()
    
        # Déboguer l'historique pour comprendre sa structure
        logging.info(f"Structure de l'historique: {self.history_manager.message_history}")
    
        # Parcourir les messages de l'historique de façon sécurisée
        try:
            for message in self.history_manager.message_history:
                # Vérifier que message est un dictionnaire
                if not isinstance(message, dict):
                    logging.error(f"Message dans un format inattendu: {message}")
                    continue
                
                # Extraire le rôle et le contenu de façon sécurisée
                role = message.get("role")
                content = message.get("content")
            
                # Vérifier que role et content sont bien définis
                if role is None or content is None:
                    logging.error(f"Message incomplet (manque role ou content): {message}")
                    continue
            
                # Afficher dans la fenêtre principale si c'est un message utilisateur ou de l'assistant principal
                if role in ["user", "assistant"]:
                    prefix = "Vous: " if role == "user" else "Assistant: "
                    self.chat_box.config(state=tk.NORMAL)
                    self.chat_box.insert(END, f"{prefix}{content}\n\n", role)

                # Afficher dans la deuxième fenêtre si c'est un message de l'assistant 2 et que la fenêtre existe
                elif role == "assistant2" and hasattr(self, 'chat_box2'):
                    self.chat_box2.config(state=tk.NORMAL)
                    self.chat_box2.insert(END, f"Assistant 2: {content}\n\n", "assistant")
                    self.chat_box2.config(state=tk.NORMAL)
            
                # Log pour les rôles non reconnus
                else:
                    logging.warning(f"Role non reconnu dans l'historique: {role}")
        
                # Défiler jusqu'au bas des fenêtres
                self.chat_box.yview(END)
                if hasattr(self, 'chat_box2'):
                    self.chat_box2.yview(END)
        
                # S'assurer que la fenêtre principale reste en mode NORMAL pour permettre la sélection
                self.chat_box.config(state=tk.NORMAL)
        
        except Exception as e:
            logging.error(f"Erreur lors du chargement de l'historique: {e}")
            messagebox.showerror("Erreur", f"Erreur lors du chargement de l'historique: {e}")

    def insert_newline(self, event):
        self.user_input.insert("insert", "\n")
        return "break"

    def attach_file(self):
        """Permet à l'utilisateur de joindre un fichier à la conversation"""
        # Ouvrir la boîte de dialogue pour sélectionner un fichier
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier",
            filetypes=[
                ("Tous les fichiers", "*.*"),
                ("Images", "*.jpg *.jpeg *.png *.gif *.bmp"),
                ("Documents", "*.pdf *.doc *.docx *.txt"),
                ("Fichiers Excel", "*.xlsx *.xls"),
                ("Fichiers CSV", "*.csv"),
                ("Fichiers Python", "*.py"),
            ],
        )

        if not file_path:
            return

        # Vérifier la taille du fichier (20MB max)
        try:
            file_size = os.path.getsize(file_path)
            if file_size > 20 * 1024 * 1024:  # 20MB
                messagebox.showerror(
                    "Fichier trop volumineux",
                    "Le fichier sélectionné dépasse la limite de 20MB."
                )
                return

            # Créer un identifiant unique pour le fichier
            file_id = len(self.attachments)
            file_name = os.path.basename(file_path)
            file_type = mimetypes.guess_type(file_path)[0] or "application/octet-stream"
            
          
            # Stocker les informations du fichier
            self.attachments[file_id] = {
                "path": file_path,
                "name": file_name,
                "type": file_type,
                "preview": None
            }
           
            # Afficher le fichier joint dans l'interface avec indicateur de chargement
            frame = self.display_attachment(file_id, file_name, loading=True)
            
            # Lors de l'utilisation du mode Eden, on s'assure d'avoir un bon traitement des fichiers Excel
            if file_path.lower().endswith(('.xlsx', '.xls')):
                logging.info(f"Fichier Excel joint: {file_path}")
                # Eden et l'analyse de fichiers Excel se font bien dans display_attachment maintenant
            
        except Exception as e:
            logging.error(f"Erreur lors de l'attachement du fichier: {e}")
            messagebox.showerror(
                "Erreur",
                f"Impossible d'attacher le fichier: {e}"
            )

    def display_attachment(self, file_id, file_name, loading=False):
        """Affiche une pièce jointe dans l'interface"""
        try:
            # Créer un cadre pour contenir la pièce jointe et son nom
            frame = Frame(self.attachments_frame, bd=1, relief=SOLID, padx=5, pady=5)
            frame.pack(side=LEFT, padx=5, pady=5)
            
            # Étiquette de chargement/traitement
            loading_label = Label(frame, text="Traitement en cours...", fg="blue")
            
            if loading:
                loading_label.pack(pady=5)
            
            # Détecter le type de fichier pour l'icône appropriée
            file_path = self.attachments[file_id]["path"]
            file_ext = os.path.splitext(file_name)[1].lower()
            
            # Image de prévisualisation par défaut
            preview_image = None
            
            # Dimensions maximales pour la prévisualisation
            max_width = 100
            max_height = 100
            
            # Gérer les fichiers image
            if file_ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                try:
                    # Charger l'image si PIL est disponible
                    if Image:
                        img = Image.open(file_path)
                        
                        # Redimensionner l'image en gardant le ratio
                        width, height = img.size
                        if width > max_width or height > max_height:
                            ratio = min(max_width/width, max_height/height)
                            new_size = (int(width * ratio), int(height * ratio))
                            img = img.resize(new_size, Image.LANCZOS)
                        
                        # Convertir en PhotoImage pour tkinter
                        preview_image = ImageTk.PhotoImage(img)
                        
                        # Créer un bouton avec l'image
                        img_btn = Button(frame, image=preview_image, 
                                        command=lambda fp=file_path: os.startfile(fp))
                        img_btn.image = preview_image  # Garder une référence
                        img_btn.pack(pady=5)
                except Exception as e:
                    logging.error(f"Erreur lors de la prévisualisation de l'image: {e}")
                    # Fallback au nom du fichier uniquement
            
            # Gérer les fichiers Excel
            elif file_ext in ['.xlsx', '.xls']:
                # Si l'extraction est toujours en cours, afficher uniquement le chargement
                if loading:
                    # Démarrer le thread pour lire le fichier Excel en arrière-plan
                    def process_excel_background():
                        try:
                            excel_preview = self.extract_excel_preview(file_path)
                            # Mettre à jour l'UI dans le thread principal
                            self.root.after(10, lambda: self.update_attachment_display(
                                file_id, frame, loading_label, None, excel_preview
                            ))
                        except Exception as e:
                            logging.error(f"Erreur lors de l'extraction Excel: {e}")
                            # Notifier l'UI de l'erreur
                            self.root.after(10, lambda: self.update_attachment_display(
                                file_id, frame, loading_label, str(e)
                            ))
                    
                    # Lancer le traitement en arrière-plan
                    threading.Thread(target=process_excel_background, daemon=True).start()
                    
                # Mais afficher aussi le lien vers le fichier
                file_link = Label(
                    frame, 
                    text=f"{file_name}", 
                    fg="blue", 
                    cursor="hand2", 
                    font=("Arial", 8, "underline")
                )
                file_link.pack(pady=2)
                file_link.bind("<Button-1>", lambda e, fp=file_path: os.startfile(fp))
            
            # Pour les autres types de fichiers, afficher simplement le nom
            else:
                file_link = Label(
                    frame, 
                    text=f"{file_name}", 
                    fg="blue", 
                    cursor="hand2",
                    font=("Arial", 8, "underline")
                )
                file_link.pack(pady=2)
                file_link.bind("<Button-1>", lambda e, fp=file_path: os.startfile(fp))
            
            # Ajouter un bouton pour supprimer la pièce jointe
            delete_btn = Button(
                frame, 
                text="✕", 
                command=lambda fid=file_id, f=frame: self.remove_attachment(fid, f),
                font=("Arial", 8), 
                padx=2, 
                pady=0
            )
            delete_btn.pack(anchor=NE)
            
            return frame
            
        except Exception as e:
            logging.error(f"Erreur lors de l'affichage de la pièce jointe: {e}")
            messagebox.showerror("Erreur", f"Impossible d'afficher la pièce jointe: {e}")
            return None
            
    def update_attachment_display(self, file_id, frame, loading_label, error=None, excel_preview=None):
        """Met à jour l'affichage d'une pièce jointe avec les données Excel si disponibles"""
        try:
            # Supprimer l'indicateur de chargement
            loading_label.destroy()
            
            # Si une erreur s'est produite
            if error:
                error_label = Label(frame, text="Erreur de lecture", fg="red", font=("Arial", 8))
                error_label.pack(pady=1)
                return
                
            # Si aucune donnée Excel n'est fournie, ne rien faire de plus
            if not excel_preview or "error" in excel_preview:
                return
                
            # Nombre de feuilles
            sheets = excel_preview.get("sheets", [])
            if not sheets:
                return
                
            # Afficher un résumé du contenu de la première feuille
            if sheets and excel_preview.get("first_rows") and sheets[0] in excel_preview["first_rows"]:
                sheet_name = sheets[0]
                sheet_data = excel_preview["first_rows"][sheet_name]
                headers = sheet_data.get("headers", [])
                
                # Créer un label avec un résumé du contenu
                if headers:
                    headers_text = ", ".join(headers[:3])
                    if len(headers) > 3:
                        headers_text += "..."
                    
                    # Afficher un aperçu des colonnes
                    cols_label = Label(
                        frame, 
                        text=f"Colonnes: {headers_text}", 
                        fg="dark green",
                        font=("Arial", 8),
                        wraplength=200
                    )
                    cols_label.pack(anchor=W, pady=1)
                    
                    # Afficher un aperçu du nombre de lignes
                    rows_label = Label(
                        frame, 
                        text=f"Lignes: {sheet_data.get('total_rows', 'Inconnu')}", 
                        fg="dark green",
                        font=("Arial", 8)
                    )
                    rows_label.pack(anchor=W, pady=1)
                    
                    # Si nous avons plusieurs feuilles, indiquer le nombre
                    if len(sheets) > 1:
                        sheets_label = Label(
                            frame, 
                            text=f"{len(sheets)} feuilles", 
                            font=("Arial", 8)
                        )
                        sheets_label.pack(anchor=W, pady=1)
                
        except Exception as e:
            logging.error(f"Erreur lors de la mise à jour de l'affichage Excel: {e}")
            # Ne pas afficher d'erreur à l'utilisateur - juste un log


    def remove_attachment(self, file_id, frame):
        """Supprime une pièce jointe"""
        try:
            # Enregistrer que le fichier est en cours de suppression
            logging.info(f"Suppression de la pièce jointe {file_id}")
            
            # Si c'est un fichier Excel, on peut capturer les données avant de supprimer
            attachment = self.attachments.get(file_id)
            if attachment and 'path' in attachment:
                file_path = attachment['path']
                if file_path.lower().endswith(('.xlsx', '.xls')):
                    logging.info(f"Suppression d'un fichier Excel: {file_path}")
            
            # Supprimer les références et la frame
            if file_id in self.attachments:
                del self.attachments[file_id]
            
            if frame and frame.winfo_exists():
                frame.destroy()
                
            logging.info(f"Pièce jointe {file_id} supprimée avec succès")
        except Exception as e:
            logging.error(f"Erreur lors de la suppression de la pièce jointe: {e}")
            messagebox.showerror("Erreur", f"Impossible de supprimer la pièce jointe: {e}")

    def append_message(self, role, message):
            prefix = "Vous: " if role == "user" else "Assistant: "
            content = f"{prefix}{message}\n\n"
    
            if role == "user":
                self.chat_box.config(state=tk.NORMAL)
                self.chat_box.insert(END, content, role)
                self.chat_box.yview(END)
                self.chat_box.config(state=tk.NORMAL)
    
                # Afficher également dans la deuxième fenêtre si elle existe
                if hasattr(self, 'chat_box2') and self.config.get("dual_chat", False):
                    self.chat_box2.config(state=tk.NORMAL)
                    self.chat_box2.insert(END, f"Vous: {message}\n\n", "user")
                    self.chat_box2.yview(END)
                    self.chat_box2.config(state=tk.NORMAL)   

            # Afficher les messages de l'assistant principal uniquement dans la première fenêtre
            elif role == "assistant":
                  self.chat_box.config(state=tk.NORMAL)
                  self.chat_box.insert(END, content, role)
                  self.chat_box.yview(END)
                  self.chat_box.config(state=tk.NORMAL)
            # Afficher les messages du deuxième assistant uniquement dans la deuxième fenêtre
            elif role == "assistant2" and hasattr(self, 'chat_box2'):
                self.chat_box2.config(state=tk.NORMAL)
                self.chat_box2.insert(END, f"Assistant 2: {message}\n\n", "assistant")
                self.chat_box2.yview(END)
                self.chat_box2.config(state=tk.NORMAL)

            self.update_selection_colors()
    
    def fetch_and_display_response(self, user_message):
        try:
            # Créer un historique filtré pour le premier assistant (sans les messages "assistant2")
            primary_history = []
            for msg in self.history_manager.message_history:
                if msg.get("role") in ["user", "assistant"]:
                    primary_history.append(msg)

            # Log l'historique pour déboguer
            logging.info(f"Envoi de requête avec {len(primary_history)} messages d'historique")
        
            # Vérifier et rafraîchir la configuration du proxy
            proxy_config = {
                "proxy_enabled": self.config.get("proxy_enabled", False),
                "proxy_host": self.config.get("proxy_host", ""),
                "proxy_port": self.config.get("proxy_port", ""),
                "proxy_username": self.config.get("proxy_username", ""),
                "proxy_password": self.config.get("proxy_password", ""),
            }

            # Pour le premier assistant, utiliser Eden ou AIClient selon la configuration
            if self.config.get("eden_mode", False):
                # Mise à jour des proxies pour Eden
                proxy_url = None
                if proxy_config["proxy_enabled"]:
                   username = proxy_config["proxy_username"]
                   password = proxy_config["proxy_password"]
                   host = proxy_config["proxy_host"]
                   port = proxy_config["proxy_port"]

                   if host and port:
                        if username and password:
                            encoded_password = urllib.parse.quote(password)
                            proxy_url = f"http://{username}:{encoded_password}@{host}:{port}"
                        else:
                            proxy_url = f"http://{host}:{port}"

                        self.ai_client.proxies = {
                            'http': proxy_url,
                            'https': proxy_url
                        }


                assistant_response = self.ai_client.chat_with_ai(user_message, primary_history)

            if self.config["response_mode"] == "typing":
                self.root.after(0, self.start_typing_effect, assistant_response)

            else:
                self.root.after(0, self.display_assistant_response, assistant_response)
   
            # Fetch response for dual chat if enabled - toujours utiliser AIClient pour le dual chat
            if self.config.get("dual_chat", False):
                try:
                    # Créer un historique spécifique pour le deuxième assistant
                    dual_chat_history = []
                    for msg in self.history_manager.message_history:
                        if msg.get("role") == "user":
                            dual_chat_history.append(msg)
                        elif msg.get("role") == "assistant2":
                            dual_chat_history.append({"role": "assistant", "content": msg.get("content")})

                    # Créer un client AI temporaire pour le dual chat
                    proxy_config = {
                        "proxy_enabled": self.config.get("proxy_enabled", False),
                        "proxy_host": self.config.get("proxy_host", ""),
                        "proxy_port": self.config.get("proxy_port", ""),
                        "proxy_username": self.config.get("proxy_username", ""),
                        "proxy_password": self.config.get("proxy_password", ""),
                    }
                    
                    temp_client = AIClient(
                        self.config["api_key"],
                        Config.DEFAULT_API_URL,  # Toujours utiliser l'URL par défaut pour le dual chat
                        self.dual_chat_model,
                        self.config["timeout"],
                        proxy_config
                    )

                    # Appeler le deuxième assistant avec son propre historique
                    dual_assistant_response = temp_client.chat_with_ai(
                        user_message, dual_chat_history, model=self.dual_chat_model
                    )
                    self.root.after(0, self.display_dual_assistant_response, dual_assistant_response)
                except Exception as e:
                    logging.error(f"Erreur pendant la récupération de la réponse du second assistant: {e}")
                    if hasattr(self, 'chat_box2'):
                        self.chat_box2.config(state=tk.NORMAL)
                        self.chat_box2.insert(END, f"Erreur: {str(e)}\n\n", "assistant")
                        self.chat_box2.config(state=tk.NORMAL)
        except Exception as e:
            logging.error(f"Erreur pendant la récupération de la réponse: {e}")
            self.root.after(
               0, self.display_assistant_response, f"Une erreur s'est produite: {e}"
            )
    


    def display_dual_assistant_response(self, assistant_response):
        if hasattr(self, 'chat_box2'):
            self.chat_box2.config(state=tk.NORMAL)
            self.chat_box2.insert(END, f"Assistant 2: {assistant_response}\n\n", "assistant")
            self.chat_box2.yview(END)
            self.chat_box2.config(state=tk.NORMAL)
        self.history_manager.add_message("assistant2", assistant_response)
        self.stop_loading()

    def start_typing_effect(self, response_text):
        self.is_typing = True
        self.chat_box.insert(END, "Assistant: ", "assistant")
        if self.bold_text_enabled:
            self.chat_box.tag_add("bold", END + "-1c", END)  # Applique le gras au "Assistant: "
        typing_thread = threading.Thread(
            target=self.typing_effect, args=(response_text,)
        )
        typing_thread.daemon = True
        typing_thread.start()

    def typing_effect(self, response_text):
        delay = self.config["typing_delay"]
        for char in response_text:
            if not self.is_typing:
                break
            # Assurer NORMAL avant l'insertion
            self.chat_box.config(state=tk.NORMAL)
            self.chat_box.insert(END, char, "assistant")
            self.chat_box.yview(END)
            self.root.update_idletasks()
            time.sleep(delay)
        # Assurer NORMAL avant l'insertion
        self.chat_box.config(state=tk.NORMAL)
        self.chat_box.insert(END, "\n\n", "assistant") 
        lang, code = CodeHandler.extract_code(response_text)
        if code:
          self.root.after(0, self.process_code_in_response, lang, code)
        self.history_manager.add_message("assistant", response_text)
        self.stop_loading()
        self.is_typing = False


    def process_code_in_response(self, lang, code):
        if not code:
            return
        theme = Config.THEMES.get(self.config["theme"], Config.THEMES["light"])

        # Créer un cadre pour les boutons de code
        code_buttons_frame = tk.Frame(self.button_frame, bg=theme["bg"])
        code_buttons_frame.grid(
            row=1,
            column=0,
            columnspan=4,  # Mis à jour pour couvrir les 4 colonnes
            padx=5,
            pady=5,
            sticky="ew",
        )

        # Bouton pour copier le code
        copy_button = Button(
            code_buttons_frame,
            text=f"Copier le code {lang if lang else ''}",
            command=lambda: CodeHandler.copy_to_clipboard(code),
            bg=theme["button_bg"],
            fg=theme["button_fg"],
        )
        copy_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Bouton pour exporter le code
        export_button = Button(
            code_buttons_frame,
            text=f"Exporter en {lang if lang else 'fichier'}",
            command=lambda: self.export_code(lang, code),
            bg=theme["button_bg"],
            fg=theme["button_fg"],
        )
        export_button.pack(side=tk.RIGHT, padx=5, fill=tk.X, expand=True)

        self.show_code_window(lang, code)
        # Augmenté à 30 secondes
        self.root.after(30000, code_buttons_frame.destroy)

    def add_email_copy_button(self):
        theme = Config.THEMES.get(self.config["theme"], Config.THEMES["light"])
        
        # Créer un cadre pour le bouton de copie d'email
        email_button_frame = tk.Frame(self.button_frame, bg=theme["bg"])
        email_button_frame.grid(
            row=1,
            column=0,
            columnspan=4,
            padx=5,
            pady=5,
            sticky="ew",
        )
        
        # Bouton pour copier l'email
        copy_button = Button(
            email_button_frame,
            text="Copier le contenu de l'email",
            command=self.detect_and_copy_email,
            bg=theme["button_bg"],
            fg=theme["button_fg"],
        )
        copy_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # Suppression automatique après un délai
        self.root.after(30000, email_button_frame.destroy)

    def show_code_window(self, language, code):
        code_window = Toplevel(self.root)
        code_window.title(f"Code {language if language else ''}")
        code_window.geometry("600x400")
        code_window.columnconfigure(0, weight=1)
        code_window.rowconfigure(0, weight=1)

        code_text = Text(code_window, wrap=tk.WORD, font=("Courier New", 10))
        code_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        code_scrollbar = Scrollbar(code_window, command=code_text.yview)
        code_scrollbar.grid(row=0, column=1, sticky="ns")
        code_text["yscrollcommand"] = code_scrollbar.set
        code_text.insert(END, code)

        # Frame pour les boutons
        button_frame = tk.Frame(code_window)
        button_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        # Bouton pour copier
        copy_button = Button(
            button_frame,
            text="Copier dans le presse-papiers",
            command=lambda: CodeHandler.copy_to_clipboard(code),
            width=20, # <-- Ajout de width
        )
        copy_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Bouton pour exporter
        export_button = Button(
            button_frame,
            text=f"Exporter en {language if language else 'fichier'}",
            command=lambda: self.export_code(language, code),
            width=20, # <-- Ajout de width
        )
        export_button.pack(side=tk.RIGHT, padx=5, fill=tk.X, expand=True)

        theme = Config.THEMES.get(self.config["theme"], Config.THEMES["light"])
        code_text.config(
            bg=theme["input_bg"],
            fg=theme["input_fg"],
            insertbackground=theme["insertbackground"],
        )
        copy_button.config(bg=theme["button_bg"], fg=theme["button_fg"])
        export_button.config(bg=theme["button_bg"], fg=theme["button_fg"])
        code_window.resizable(True, True)

    def export_code(self, language, code):
        # Exporte le code dans un fichier
        if not code:
            messagebox.showerror("Erreur", "Aucun code à exporter.")
            return

        # Déterminer l'extension de fichier appropriée
        extensions = {
            "python": ".py",
            "py": ".py",
            "javascript": ".js",
            "js": ".js",
            "html": ".html",
            "css": ".css",
            "java": ".java",
            "c": ".c",
            "cpp": ".cpp",
            "c++": ".cpp",
            "csharp": ".cs",
            "cs": ".cs",
            "php": ".php",
            "ruby": ".rb",
            "go": ".go",
            "rust": ".rs",
            "swift": ".swift",
            "kotlin": ".kt",
            "typescript": ".ts",
            "ts": ".ts",
            "sql": ".sql",
            "bash": ".sh",
            "sh": ".sh",
            "powershell": ".ps1",
            "ps1": ".ps1",
            "r": ".r",
            "markdown": ".md",
            "md": ".md",
            "json": ".json",
            "xml": ".xml",
            "yaml": ".yml",
            "yml": ".yml",
        }

        # Déterminer l'extension par défaut
        if language and language.lower() in extensions:
            default_ext = extensions[language.lower()]
        else:
            default_ext = ".txt"

        # Suggérer un nom de fichier
        suggested_filename = f"code_export{default_ext}"

        # Demander où sauvegarder le fichier
        file_path = filedialog.asksaveasfilename(
            defaultextension=default_ext,
            filetypes=[
                ("Fichier " + (language if language else "texte"), "*" + default_ext),
                ("Tous les fichiers", "*.*"),
            ],
            initialfile=suggested_filename,
        )

        if not file_path:
            return  # L'utilisateur a annulé

        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(code)
            messagebox.showinfo(
                "Exportation réussie",
                f"Le code a été exporté avec succès dans {file_path}",
            )
        except Exception as e:
            messagebox.showerror(
                "Erreur d'exportation", f"Erreur lors de l'exportation du code: {e}"
            )

    def start_loading(self):
        self.loading_label.config(text="Génération de la réponse...")
        self.loading_label.grid(row=4, column=0, padx=10, pady=5, sticky="w")
        self.cancel_button.config(state=tk.NORMAL)
        self.root.update_idletasks()

    def stop_loading(self):
        self.loading_label.grid_forget()
        self.send_button.config(state=tk.NORMAL)
        self.cancel_button.config(state=tk.NORMAL)
        self.root.update_idletasks()

    def cancel_generation(self):
        if self.is_typing:
            self.is_typing = False
        self.ai_client.cancel_request()
        self.stop_loading()
        self.chat_box.insert(END, "[Génération annulée]\n\n", "assistant")
        self.chat_box.yview(END)

    def copy_selected_text(self):
        try:
            selected_text = self.chat_box.get("sel.first", "sel.last")
            if selected_text:
                CodeHandler.copy_to_clipboard(selected_text)
        except tk.TclError:
            pass

    def copy_code_from_chat(self):
        chat_content = self.chat_box.get("1.0", END)
        lang, code = CodeHandler.extract_code(chat_content)
        if code:
            CodeHandler.copy_to_clipboard(code)
        else:
            messagebox.showinfo("Information", "Aucun code trouvé dans le chat.")

    def copy_excel_from_chat(self):
        """Copie les formules Excel ou le code VBA trouvé dans le chat"""
        chat_content = self.chat_box.get("1.0", END)
        excel_type, excel_content = CodeHandler.extract_excel_formula(chat_content)
        
        if excel_content:
            CodeHandler.copy_to_clipboard(excel_content)
            messagebox.showinfo("Information", f"Formule/code {excel_type} copié dans le presse-papiers!")
        else:
            messagebox.showinfo("Information", "Aucune formule Excel ou code VBA trouvé dans le chat.")

    def search_text(self):
        search_window = Toplevel(self.root)
        search_window.title("Rechercher")
        search_window.geometry("300x100")
        search_window.resizable(False, False)
        Label(search_window, text="Texte à rechercher:").pack(pady=5)
        search_entry = Entry(search_window, width=30)
        search_entry.pack(pady=5)
        search_entry.focus_set()

        def search():
            query = search_entry.get()
            if not query:
                return
            self.chat_box.tag_remove("search", "1.0", END)
            start_pos = "1.0"
            count = 0
            while True:
                start_pos = self.chat_box.search(
                    query, start_pos, stopindex=END, nocase=True
                )
                if not start_pos:
                    break
                end_pos = f"{start_pos}+{len(query)}c"
                self.chat_box.tag_add("search", start_pos, end_pos)
                self.chat_box.tag_config(
                    "search", background="yellow", foreground="black"
                )
                start_pos = end_pos
                count += 1
            if count:
                messagebox.showinfo("Recherche", f"{count} occurrence(s) trouvée(s)")
            else:
                messagebox.showinfo("Recherche", "Aucune occurrence trouvée")
            search_window.destroy()

        Button(search_window, text="Rechercher", command=search).pack(pady=5)
        search_window.bind("<Return>", lambda e: search())

    def open_config_window(self):
        config_window = Toplevel(self.root)
        config_window.title("Configuration")
        config_window.geometry("800x700")  # Augmentation de la taille de la fenêtre
        config_window.minsize(800, 700)    # Taille minimale
        config_window.resizable(True, True)

        # Configurer le redimensionnement
        config_window.columnconfigure(0, weight=1)
        config_window.rowconfigure(0, weight=1)

        notebook = ttk.Notebook(config_window)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        # Onglet API
        api_frame = ttk.Frame(notebook)
        notebook.add(api_frame, text="API")
        
        # Configurer le redimensionnement de l'onglet API
        api_frame.columnconfigure(1, weight=1)  # La colonne des entrées s'étend
        api_frame.columnconfigure(2, weight=0)  # La colonne du bouton reste fixe

        ttk.Label(api_frame, text="Clé API:").grid(
            row=0, column=0, padx=10, pady=5, sticky=W
        )
        api_key_entry = ttk.Entry(api_frame)
        api_key_entry.grid(row=0, column=1, columnspan=2, padx=10, pady=5, sticky=EW)
        api_key_entry.insert(0, self.config["api_key"])

        ttk.Label(api_frame, text="URL API:").grid(
            row=1, column=0, padx=10, pady=5, sticky=W
        )
        api_url_entry = ttk.Entry(api_frame)
        api_url_entry.grid(row=1, column=1, columnspan=2, padx=10, pady=5, sticky=EW)
        api_url_entry.insert(0, self.config["api_url"])
        
        ttk.Label(api_frame, text="Modèle:").grid(
            row=2, column=0, padx=10, pady=5, sticky=W
        )

        # Remplacer l'Entry par un Combobox pour le modèle avec une largeur plus grande
        model_var = tk.StringVar(value=self.config["model"])
        model_combo = ttk.Combobox(api_frame, textvariable=model_var)
        model_combo.grid(row=2, column=1, padx=10, pady=5, sticky=EW)

        # Choisir la liste de modèles selon le mode (local ou en ligne)
        if self.config.get("local_mode", False):
            model_combo["values"] = Config.LOCAL_MODELS
        else:
            model_combo["values"] = Config.COMMON_MODELS

        # Permettre également l'entrée manuelle d'un modèle non listé
        model_combo.set(self.config["model"])

        # Ajouter le bouton de mise à jour des modèles
        def update_models():
            success, message = Config.update_models()
            if success:
                # Mettre à jour la liste des modèles dans le Combobox
                if not self.config.get("local_mode", False):
                    model_combo["values"] = Config.COMMON_MODELS
                messagebox.showinfo("Succès", message)
            else:
                messagebox.showerror("Erreur", message)

        update_models_button = ttk.Button(
            api_frame,
            text="Mettre à jour les modèles",
            command=update_models
        )
        update_models_button.grid(row=2, column=2, padx=5, pady=5)

        ttk.Label(api_frame, text="Timeout (secondes):").grid(
            row=3, column=0, padx=10, pady=5, sticky=W
        )
        timeout_scale = Scale(api_frame, from_=5, to=120, orient=tk.HORIZONTAL)
        timeout_scale.grid(row=3, column=1, columnspan=2, padx=10, pady=5, sticky=EW)
        timeout_scale.set(self.config["timeout"])

        ttk.Label(api_frame, text="Modèle Dual Chat:").grid(
            row=4, column=0, padx=10, pady=5, sticky=W
        )
        dual_model_var = tk.StringVar(value=self.config.get("dual_chat_model", self.config["model"]))
        dual_model_combo = ttk.Combobox(api_frame, textvariable=dual_model_var)
        dual_model_combo.grid(row=4, column=1, columnspan=2, padx=10, pady=5, sticky=EW)
        dual_model_combo["values"] = Config.COMMON_MODELS
        dual_model_combo.set(self.config.get("dual_chat_model", self.config["model"]))

        # Onglet Interface
        ui_frame = ttk.Frame(notebook)
        notebook.add(ui_frame, text="Interface")

        ttk.Label(ui_frame, text="Mode de réponse:").grid(
            row=0, column=0, padx=10, pady=5, sticky=W
        )
        response_mode_var = tk.StringVar(value=self.config["response_mode"])
        ttk.Radiobutton(
            ui_frame, text="Effet de saisie", variable=response_mode_var, value="typing"
        ).grid(row=0, column=1, padx=10, pady=5, sticky=W)
        ttk.Radiobutton(
            ui_frame, text="Instantané", variable=response_mode_var, value="instant"
        ).grid(row=0, column=2, padx=10, pady=5, sticky=W)

        ttk.Label(ui_frame, text="Délai de saisie:").grid(
            row=1, column=0, padx=10, pady=5, sticky=W
        )
        delay_scale = Scale(
            ui_frame, from_=0.001, to=0.1, resolution=0.001, orient=tk.HORIZONTAL
        )
        delay_scale.grid(row=1, column=1, columnspan=2, padx=10, pady=5, sticky=EW)
        delay_scale.set(self.config["typing_delay"])

        ttk.Label(ui_frame, text="Thème:").grid(
            row=2, column=0, padx=10, pady=5, sticky=W
        )
        theme_var = tk.StringVar(value=self.config["theme"])
        theme_combo = ttk.Combobox(
            ui_frame, textvariable=theme_var, values=list(Config.THEMES.keys())
        )
        theme_combo.grid(row=2, column=1, columnspan=2, padx=10, pady=5, sticky=EW)

        ttk.Label(ui_frame, text="Largeur fenêtre:").grid(
            row=3, column=0, padx=10, pady=5, sticky=W
        )
        width_entry = ttk.Entry(ui_frame, width=10)
        width_entry.grid(row=3, column=1, padx=10, pady=5, sticky=W)
        width_entry.insert(0, str(self.config["window_width"]))

        ttk.Label(ui_frame, text="Hauteur fenêtre:").grid(
            row=3, column=2, padx=10, pady=5, sticky=W
        )
        height_entry = ttk.Entry(ui_frame, width=10)
        height_entry.grid(row=3, column=3, padx=10, pady=5, sticky=W)
        height_entry.insert(0, str(self.config["window_height"]))

        # Onglet Local
        local_frame = ttk.Frame(notebook)
        notebook.add(local_frame, text="Local")
        cpu_only_var = tk.BooleanVar(value=self.config.get("cpu_only", False))
        ttk.Checkbutton(local_frame, text="CPU only", variable=cpu_only_var).grid(
            row=0, column=0, padx=10, pady=5, sticky=W
        )
        ttk.Label(local_frame, text="URL API locale:").grid(
            row=1, column=0, padx=10, pady=5, sticky=W
        )
        local_url_entry = ttk.Entry(local_frame, width=40)
        local_url_entry.grid(row=1, column=1, padx=10, pady=5, sticky=EW)
        local_url_entry.insert(
            0, self.config.get("local_api_url", Config.DEFAULT_LOCAL_API_URL)
        )
        ttk.Label(local_frame, text="Modèle local:").grid(
            row=2, column=0, padx=10, pady=5, sticky=W
        )
        local_model_entry = ttk.Entry(local_frame, width=40)
        local_model_entry.grid(row=2, column=1, padx=10, pady=5, sticky=EW)
        local_model_entry.insert(
            0, self.config.get("local_model", Config.DEFAULT_LOCAL_MODEL)
        )

    # Ajouter un nouvel onglet Proxy
        proxy_frame = ttk.Frame(notebook)
        notebook.add(proxy_frame, text="Proxy")
        
        # Checkbox pour activer/désactiver le proxy
        proxy_enabled_var = tk.BooleanVar(value=self.config.get("proxy_enabled", False))
        proxy_checkbox = ttk.Checkbutton(proxy_frame, text="Activer le proxy", variable=proxy_enabled_var)
        proxy_checkbox.grid(row=0, column=0, columnspan=2, padx=10, pady=5, sticky=W)
        
        # Champs pour les paramètres du proxy
        ttk.Label(proxy_frame, text="Hôte:").grid(row=1, column=0, padx=10, pady=5, sticky=W)
        proxy_host_entry = ttk.Entry(proxy_frame, width=40)
        proxy_host_entry.grid(row=1, column=1, padx=10, pady=5, sticky=EW)
        proxy_host_entry.insert(0, self.config.get("proxy_host", ""))
        
        ttk.Label(proxy_frame, text="Port:").grid(row=2, column=0, padx=10, pady=5, sticky=W)
        proxy_port_entry = ttk.Entry(proxy_frame, width=40)
        proxy_port_entry.grid(row=2, column=1, padx=10, pady=5, sticky=EW)
        proxy_port_entry.insert(0, self.config.get("proxy_port", ""))
        
        ttk.Label(proxy_frame, text="Nom d'utilisateur:").grid(row=3, column=0, padx=10, pady=5, sticky=W)
        proxy_username_entry = ttk.Entry(proxy_frame, width=40)
        proxy_username_entry.grid(row=3, column=1, padx=10, pady=5, sticky=EW)
        proxy_username_entry.insert(0, self.config.get("proxy_username", ""))
        
        ttk.Label(proxy_frame, text="Mot de passe:").grid(row=4, column=0, padx=10, pady=5, sticky=W)
        proxy_password_entry = ttk.Entry(proxy_frame, width=40, show="*")
        proxy_password_entry.grid(row=4, column=1, padx=10, pady=5, sticky=EW)
        proxy_password_entry.insert(0, self.config.get("proxy_password", ""))
        
        # Fonction de gestion de l'état activé/désactivé pour les champs du proxy
        def update_proxy_fields_state(*args):
            state = "normal" if proxy_enabled_var.get() else "disabled"
            proxy_host_entry.config(state=state)
            proxy_port_entry.config(state=state)
            proxy_username_entry.config(state=state)
            proxy_password_entry.config(state=state)
        
        # Appliquer l'état initial
        proxy_enabled_var.trace_add("write", update_proxy_fields_state)
        update_proxy_fields_state()

        # Ajouter un nouvel onglet Eden
        eden_frame = ttk.Frame(notebook)
        notebook.add(eden_frame, text="Eden")
        
        # Checkbox pour activer/désactiver le mode Eden
        eden_mode_var = tk.BooleanVar(value=self.config.get("eden_mode", False))
        eden_checkbox = ttk.Checkbutton(eden_frame, text="Activer Eden AI", variable=eden_mode_var)
        eden_checkbox.grid(row=0, column=0, columnspan=2, padx=10, pady=5, sticky=W)
        
        # URL de l'API Eden
        ttk.Label(eden_frame, text="URL API Eden:").grid(row=1, column=0, padx=10, pady=5, sticky=W)
        eden_url_entry = ttk.Entry(eden_frame, width=40)
        eden_url_entry.grid(row=1, column=1, padx=10, pady=5, sticky=EW)
        eden_url_entry.insert(0, self.config.get("eden_api_url", Config.DEFAULT_EDEN_API_URL))
        
        # Clé API Eden
        ttk.Label(eden_frame, text="Clé API Eden:").grid(row=2, column=0, padx=10, pady=5, sticky=W)
        eden_key_entry = ttk.Entry(eden_frame, width=40)
        eden_key_entry.grid(row=2, column=1, padx=10, pady=5, sticky=EW)
        eden_key_entry.insert(0, self.config.get("eden_api_key", Config.DEFAULT_EDEN_API_KEY))
        
        # ID Geni Eden
        ttk.Label(eden_frame, text="ID Geni Eden:").grid(row=3, column=0, padx=10, pady=5, sticky=W)
        eden_geni_entry = ttk.Entry(eden_frame, width=40)
        eden_geni_entry.grid(row=3, column=1, padx=10, pady=5, sticky=EW)
        eden_geni_entry.insert(0, self.config.get("eden_geni_id", Config.DEFAULT_EDEN_GENI_ID))
        
        # Option pour la vérification internet
        eden_internet_verify_var = tk.BooleanVar(value=self.config.get("eden_internet_verify", Config.DEFAULT_EDEN_INTERNET_VERIFY))
        eden_internet_verify_checkbox = ttk.Checkbutton(
            eden_frame, 
            text="Activer la vérification des informations sur internet", 
            variable=eden_internet_verify_var
        )
        eden_internet_verify_checkbox.grid(row=4, column=0, columnspan=2, padx=10, pady=5, sticky=W)
        ttk.Label(eden_frame, text="Note: Cette option permet à Eden de rechercher des informations sur internet pour améliorer ses réponses.").grid(
            row=5, column=0, columnspan=2, padx=10, pady=5, sticky=W)

        # Ajouter un nouvel onglet OCR
        ocr_frame = ttk.Frame(notebook)
        notebook.add(ocr_frame, text="OCR")
        
        # Checkbox pour activer/désactiver l'OCR
        ocr_enabled_var = tk.BooleanVar(value=self.config.get("ocr_enabled", Config.DEFAULT_OCR_ENABLED))
        ocr_checkbox = ttk.Checkbutton(ocr_frame, text="Activer l'OCR pour images", variable=ocr_enabled_var)
        ocr_checkbox.grid(row=0, column=0, columnspan=2, padx=10, pady=5, sticky=W)
        
        # Sélection du mode OCR
        ttk.Label(ocr_frame, text="Mode OCR:").grid(row=1, column=0, padx=10, pady=5, sticky=W)
        ocr_mode_var = tk.StringVar(value=self.config.get("ocr_mode", Config.DEFAULT_OCR_MODE))
        ocr_mode_combo = ttk.Combobox(ocr_frame, textvariable=ocr_mode_var, values=["online", "local"])
        ocr_mode_combo.grid(row=1, column=1, padx=10, pady=5, sticky=EW)
        
        # Clé API OCR (pour OCR.space)
        ttk.Label(ocr_frame, text="Clé API OCR.space:").grid(row=2, column=0, padx=10, pady=5, sticky=W)
        ocr_api_key_entry = ttk.Entry(ocr_frame, width=40)
        ocr_api_key_entry.grid(row=2, column=1, padx=10, pady=5, sticky=EW)
        ocr_api_key_entry.insert(0, self.config.get("ocr_api_key", Config.DEFAULT_OCR_API_KEY))
        
        # Note explicative sur l'OCR gratuit
        ttk.Label(ocr_frame, text="Note: La clé 'helloworld' permet une utilisation gratuite limitée.").grid(
            row=3, column=0, columnspan=2, padx=10, pady=5, sticky=W)
        
        # Langues OCR
        ttk.Label(ocr_frame, text="Langues OCR:").grid(row=4, column=0, padx=10, pady=5, sticky=W)
        ocr_lang_entry = ttk.Entry(ocr_frame, width=40)
        ocr_lang_entry.grid(row=4, column=1, padx=10, pady=5, sticky=EW)
        ocr_lang_entry.insert(0, self.config.get("ocr_lang", Config.DEFAULT_OCR_LANG))
        
        # Note explicative sur les langues
        ttk.Label(ocr_frame, text="Format: 'fra+eng' pour français et anglais.").grid(
            row=5, column=0, columnspan=2, padx=10, pady=5, sticky=W)
        
        # Fonction de test OCR
        def test_ocr():
            from Chat import OCRHandler  # Importer explicitement la classe OCRHandler
            
            file_path = filedialog.askopenfilename(
                title="Sélectionner une image pour tester l'OCR",
                filetypes=[("Images", "*.png *.jpg *.jpeg *.gif *.bmp")]
            )
            if file_path:
                try:
                    # Créer une fenêtre de progression
                    progress_window = Toplevel(config_window)
                    progress_window.title("Test OCR en cours")
                    progress_window.geometry("300x100")
                    progress_window.resizable(False, False)
                    progress_window.transient(config_window)
                    progress_window.grab_set()
                    
                    Label(progress_window, text="Extraction du texte en cours...").pack(pady=10)
                    progress_bar = ttk.Progressbar(progress_window, mode="indeterminate")
                    progress_bar.pack(fill=tk.X, padx=20)
                    progress_bar.start()
                    
                    # Fonction pour exécuter l'OCR en arrière-plan
                    def run_ocr():
                        try:
                            # Utiliser les valeurs actuelles des champs
                            use_online = ocr_mode_var.get() == "online"
                            api_key = ocr_api_key_entry.get()
                            lang = ocr_lang_entry.get()

                            # Récupérer les paramètres du proxy depuis les champs de la configuration
                            proxy_config = {
                            "proxy_enabled": proxy_enabled_var.get(),
                            "proxy_host": proxy_host_entry.get(),
                            "proxy_port": proxy_port_entry.get(),
                            "proxy_username": proxy_username_entry.get(),
                            "proxy_password": proxy_password_entry.get(),
                            }

                            # Activer le journal détaillé
                            logging.getLogger().setLevel(logging.DEBUG)
                            
                            # Récupérer les informations sur l'image
                            img_info = ""
                            try:
                                with Image.open(file_path) as img:
                                    img_info = f"Format: {img.format}, Mode: {img.mode}, Taille: {img.size[0]}x{img.size[1]}"
                            except Exception as e:
                                img_info = f"Erreur lors de l'ouverture de l'image: {e}"
                            
                            # Exécuter l'OCR
                            if use_online:
                                text = OCRHandler.extract_text_online(file_path, api_key=api_key, lang=lang.split('+')[0], proxy_config=proxy_config)
                            else:
                                text = OCRHandler.extract_text(file_path, lang=lang, use_online=False)
                            
                            # Afficher le résultat dans le thread principal
                            config_window.after(0, lambda: show_ocr_result(text, img_info, use_online, api_key, lang))
                        except Exception as e:
                            logging.error(f"Erreur lors du test OCR: {e}", exc_info=True)
                            config_window.after(0, lambda e=e: show_ocr_result(f"Erreur lors du test OCR: {str(e)}", "", use_online, api_key, lang))
                        finally:
                            # Rétablir le niveau de journal normal
                            logging.getLogger().setLevel(logging.INFO)
                            config_window.after(0, progress_window.destroy)
                    
                    # Fonction pour afficher le résultat
                    def show_ocr_result(text, img_info, use_online, api_key, lang):
                        result_window = Toplevel(config_window)
                        result_window.title("Résultat du test OCR")
                        result_window.geometry("600x500")
                        
                        # Cadre principal avec défilement
                        main_frame = ttk.Frame(result_window)
                        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
                        
                        # Informations sur le test
                        info_frame = ttk.LabelFrame(main_frame, text="Informations")
                        info_frame.pack(fill=tk.X, padx=5, pady=5)
                        
                        ttk.Label(info_frame, text=f"Fichier: {os.path.basename(file_path)}").pack(anchor="w", padx=5)
                        ttk.Label(info_frame, text=f"Chemin: {file_path}").pack(anchor="w", padx=5)
                        ttk.Label(info_frame, text=f"Image: {img_info}").pack(anchor="w", padx=5)
                        ttk.Label(info_frame, text=f"Mode: {'Online (OCR.space)' if use_online else 'Local (Tesseract)'}").pack(anchor="w", padx=5)
                        ttk.Label(info_frame, text=f"Langue(s): {lang}").pack(anchor="w", padx=5)
                        
                        # Texte extrait
                        text_frame = ttk.LabelFrame(main_frame, text="Texte extrait")
                        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
                        
                        result_text = scrolledtext.ScrolledText(text_frame, wrap=tk.WORD)
                        result_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
                        result_text.insert(tk.END, text if text else "Aucun texte extrait")
                        
                        # Boutons
                        button_frame = ttk.Frame(result_window)
                        button_frame.pack(fill=tk.X, padx=10, pady=10)
                        
                        # Bouton pour copier le texte
                        ttk.Button(
                            button_frame, 
                            text="Copier le texte", 
                            command=lambda: [pyperclip.copy(text), messagebox.showinfo("Copie", "Texte copié dans le presse-papiers")]
                        ).pack(side=tk.LEFT, padx=5)
                        
                        # Bouton pour essayer l'autre mode
                        ttk.Button(
                            button_frame,
                            text=f"Essayer en mode {'local' if use_online else 'en ligne'}",
                            command=lambda: [result_window.destroy(), 
                                            ocr_mode_var.set("local" if use_online else "online"),
                                            test_ocr()]
                        ).pack(side=tk.LEFT, padx=5)
                        
                        # Bouton pour fermer
                        ttk.Button(button_frame, text="Fermer", command=result_window.destroy).pack(side=tk.RIGHT, padx=5)
                    
                    # Lancer l'OCR en arrière-plan
                    threading.Thread(target=run_ocr, daemon=True).start()
                    
                except Exception as e:
                    logging.error(f"Erreur lors du test OCR: {e}", exc_info=True)
                    messagebox.showerror("Erreur OCR", f"Erreur lors du test OCR: {str(e)}")
        
        # Bouton de test OCR
        test_button = ttk.Button(ocr_frame, text="Tester l'OCR", command=test_ocr)
        test_button.grid(row=6, column=0, columnspan=2, padx=10, pady=15)

        def save_config():
            try:
                window_width = int(width_entry.get())
                window_height = int(height_entry.get())
                new_config = {
                    "api_key": api_key_entry.get(),
                    "api_url": api_url_entry.get(),
                    "model": model_var.get(),
                    "typing_delay": float(delay_scale.get()),
                    "response_mode": response_mode_var.get(),
                    "window_width": window_width,
                    "window_height": window_height,
                    "timeout": int(timeout_scale.get()),
                    "theme": theme_var.get(),
                    "cpu_only": cpu_only_var.get(),
                    "local_api_url": local_url_entry.get(),
                    "local_model": local_model_entry.get(),
                    # Ajouter les paramètres de proxy
                    "proxy_enabled": proxy_enabled_var.get(),
                    "proxy_host": proxy_host_entry.get(),
                    "proxy_port": proxy_port_entry.get(),
                    "proxy_username": proxy_username_entry.get(),
                    "proxy_password": proxy_password_entry.get(),
                    "dual_chat_model": dual_model_var.get(),
                    # Ajouter les paramètres Eden
                    "eden_mode": eden_mode_var.get(),
                    "eden_api_url": eden_url_entry.get(),
                    "eden_api_key": eden_key_entry.get(),
                    "eden_geni_id": eden_geni_entry.get(),
                    "eden_internet_verify": eden_internet_verify_var.get(),
                    # Paramètres OCR
                    "ocr_mode": ocr_mode_var.get(),
                    "ocr_api_key": ocr_api_key_entry.get(),
                    "ocr_lang": ocr_lang_entry.get(),
                    "ocr_enabled": ocr_enabled_var.get(),
                }
                self.config.update(new_config)
                Config.save(self.config)

                # Mettre à jour la variable d'instance dual_chat_model
                self.dual_chat_model = dual_model_var.get()
                
                # Créer un dictionnaire de configuration du proxy
                # IMPORTANT: Création d'un dictionnaire proxy correct et complet
                proxy_config = {
                    "proxy_enabled": self.config.get("proxy_enabled", False),
                    "proxy_host": self.config.get("proxy_host", ""),
                    "proxy_port": self.config.get("proxy_port", ""),
                    "proxy_username": self.config.get("proxy_username", ""),
                    "proxy_password": self.config.get("proxy_password", ""),
                }

                # Recréer le client AI avec les nouveaux paramètres
                if new_config["eden_mode"]:
                    self.ai_client = EdenClient(self.config["timeout"], proxy_config)  # Passer le proxy
                    logging.info("Client Eden créé avec configuration proxy")
                else:
                    self.ai_client = AIClient(
                        self.config["api_key"],
                        self.config["api_url"],
                        self.config["model"],
                        self.config["timeout"],
                        proxy_config
                    )

                self.apply_theme(self.config["theme"])
                self.root.geometry(f"{window_width}x{window_height}")
                config_window.destroy()
                messagebox.showinfo(
                    "Configuration", "Configuration sauvegardée avec succès!"
                )
                self.update_mode_indicator()
            except ValueError as e:
                messagebox.showerror("Erreur", f"Erreur de saisie: {e}")

        save_button = ttk.Button(config_window, text="Sauvegarder", command=save_config)
        save_button.pack(side=tk.BOTTOM, pady=10)

    def show_about(self):
        about_window = Toplevel(self.root)
        about_window.title("À propos")
        about_window.geometry("400x300")
        about_window.resizable(False, False)
        Label(
            about_window,
            text="Chat avec IA - Interface Améliorée par Marco LY",
            font=("Arial", 10, "bold"),
        ).pack(pady=10)
        Label(about_window, text="Version 1.0", font=("Arial", 10, "italic")).pack()
        Label(about_window, text="© 2025 - Tous droits réservés").pack(pady=20)
        Label(
            about_window,
            text="Une interface conviviale pour discuter avec les modèles d'IA.",
        ).pack(pady=5)
        Label(
            about_window, text="Compatible avec OpenRouter et de nombreuses API de LLM."
        ).pack(pady=5)
        Button(about_window, text="Fermer", command=about_window.destroy).pack(pady=20)

    def show_help(self):
        help_window = Toplevel(self.root)
        help_window.title("Aide")
        help_window.geometry("600x400")
        help_text = scrolledtext.ScrolledText(
            help_window, wrap=tk.WORD, font=("Arial", 10)
        )
        help_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        help_content = """
# Aide - Chat avec IA

## Utilisation de base
- Tapez votre message dans la zone de texte en bas.
- Appuyez sur "Envoyer" ou sur la touche "Entrée" pour envoyer le message.
- Utilisez "Ctrl+Entrée" pour insérer un saut de ligne.
- Cliquez sur "Joindre fichier" pour ajouter des documents à votre message.

## Raccourcis clavier
- Ctrl+F : Rechercher dans le chat.
- Ctrl+N : Nouvelle conversation.
- Ctrl+S : Sauvegarder l'historique.

## Fonctionnalités
- Clic droit sur le chat pour le menu contextuel.
- Les blocs de code sont détectés automatiquement.
- Possibilité de copier et d'exporter le code généré.
- Différents thèmes disponibles dans le menu "Affichage".
- Sélection du texte améliorée dans la fenêtre de chat.
- Mode gras disponible pour mettre en évidence certaines parties du texte
- Formats supportés : texte, images, PDF, Excel, CSV, etc.
- Les fichiers texte sont inclus directement dans le message.
- Limite de taille : 10 Mo par fichier.

## Configuration
- Menu Fichier -> Configuration.
- Définissez votre clé API, modèle et autres paramètres.
- Dans l'onglet "Local", vous pouvez définir l'URL et le modèle local, et activer l'option "CPU only".

## Mode local
- En mode local, le logiciel vérifie via une connexion socket que le serveur Ollama est actif.
- Un indicateur en bas vous informe si vous êtes en mode Online ou Local (avec indication "CPU only" si activé).

## Proxy
- Vous pouvez activer ou désactiver le proxy dans l'onglet "Proxy" de la configuration.
- Si activé, vous devez spécifier l'hôte et le port du proxy.
- Le nom d'utilisateur et le mot de passe sont optionnels.
- L'état du proxy est affiché dans la barre d'état en bas de l'application.

## Problèmes courants
- Réponses lentes : Vérifiez votre configuration et le mode utilisé.
- Fichiers non lisibles : Vérifiez l'encodage (UTF-8 recommandé).

## Besoin d'aide supplémentaire?
Contactez le support technique à support@example.com
        """
        help_text.insert(END, help_content)
        help_text.config(state="normal")
        Button(help_window, text="Fermer", command=help_window.destroy).pack(pady=10)

    def load_app_state(self):
        try:
            with open(Config.PERSISTENCE_FILE, "r") as f:
                state = json.load(f)
                # Appliquer les états sauvegardés si disponibles
                if "window_width" in state and "window_height" in state:
                    self.root.geometry(
                        f"{state['window_width']}x{state['window_height']}"
                    )
        except (FileNotFoundError, json.JSONDecodeError):
            pass

    def save_app_state(self):
        state = {
            "window_width": self.root.winfo_width(),
            "window_height": self.root.winfo_height(),
        }
        try:
            with open(Config.PERSISTENCE_FILE, "w") as f:
                json.dump(state, f, indent=4)
        except Exception as e:
            logging.error(f"Erreur lors de la sauvegarde de l'état: {e}")

    def on_closing(self):
        if messagebox.askyesno("Quitter", "Voulez-vous quitter l'application ?"):
            if self.is_typing:
                self.is_typing = False
            self.ai_client.cancel_request()
            self.save_app_state()
            Config.save(self.config)
            self.history_manager.save()
            self.root.destroy()

    # ----- Fonctions pour les graphiques -----

    def open_chart_creator(self):
        """Ouvre une fenêtre pour créer des graphiques et les intégrer dans Excel"""
        if plt is None or openpyxl is None:
            messagebox.showerror(
                "Erreur",
                "Impossible de créer des graphiques. Assurez-vous que matplotlib et openpyxl sont installés."
            )
            return
        
        chart_window = Toplevel(self.root)
        chart_window.title("Créateur de graphiques")
        chart_window.geometry("800x600")
        chart_window.minsize(600, 400)
        
        # Appliquer le thème actuel
        theme = Config.THEMES.get(self.config["theme"], Config.THEMES["light"])
        chart_window.config(bg=theme["bg"])
        
        # Cadre principal
        main_frame = tk.Frame(chart_window, bg=theme["bg"])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Cadre pour les options
        options_frame = tk.LabelFrame(main_frame, text="Options du graphique", bg=theme["bg"], fg=theme["fg"])
        options_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=5, pady=5)
        
        # Type de graphique
        tk.Label(options_frame, text="Type de graphique:", bg=theme["bg"], fg=theme["fg"]).grid(row=0, column=0, sticky=W, padx=5, pady=5)
        chart_type_var = tk.StringVar(value="line")
        chart_types = [
            ("Courbes", "line"),
            ("Barres", "bar"),
            ("Camembert", "pie"),
            ("Nuage de points", "scatter"),
            ("Histogramme", "histogram")
        ]
        
        for i, (text, value) in enumerate(chart_types):
            tk.Radiobutton(
                options_frame,
                text=text,
                variable=chart_type_var,
                value=value,
                bg=theme["bg"],
                fg=theme["fg"],
                selectcolor=theme["bg"]
            ).grid(row=i+1, column=0, sticky=W, padx=5, pady=2)
        
        # Titre du graphique
        tk.Label(options_frame, text="Titre:", bg=theme["bg"], fg=theme["fg"]).grid(row=6, column=0, sticky=W, padx=5, pady=5)
        title_entry = tk.Entry(options_frame, bg=theme["input_bg"], fg=theme["input_fg"])
        title_entry.grid(row=6, column=1, sticky=EW, padx=5, pady=5)
        title_entry.insert(0, "Mon graphique")
        
        # Étiquettes des axes
        tk.Label(options_frame, text="Étiquette X:", bg=theme["bg"], fg=theme["fg"]).grid(row=7, column=0, sticky=W, padx=5, pady=5)
        xlabel_entry = tk.Entry(options_frame, bg=theme["input_bg"], fg=theme["input_fg"])
        xlabel_entry.grid(row=7, column=1, sticky=EW, padx=5, pady=5)
        xlabel_entry.insert(0, "X")
        
        tk.Label(options_frame, text="Étiquette Y:", bg=theme["bg"], fg=theme["fg"]).grid(row=8, column=0, sticky=W, padx=5, pady=5)
        ylabel_entry = tk.Entry(options_frame, bg=theme["input_bg"], fg=theme["input_fg"])
        ylabel_entry.grid(row=8, column=1, sticky=EW, padx=5, pady=5)
        ylabel_entry.insert(0, "Y")
        
        # Cadre pour les données
        data_frame = tk.LabelFrame(main_frame, text="Données", bg=theme["bg"], fg=theme["fg"])
        data_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Zone de texte pour les données
        tk.Label(data_frame, text="Format: Une valeur par ligne ou 'x,y' pour les graphiques XY", bg=theme["bg"], fg=theme["fg"]).pack(anchor=W, padx=5, pady=5)
        data_text = scrolledtext.ScrolledText(data_frame, height=10, bg=theme["input_bg"], fg=theme["input_fg"])
        data_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Pour les camemberts, ajouter une zone pour les étiquettes
        labels_frame = tk.LabelFrame(data_frame, text="Étiquettes (pour camembert)", bg=theme["bg"], fg=theme["fg"])
        labels_frame.pack(fill=tk.X, padx=5, pady=5)
        
        labels_text = scrolledtext.ScrolledText(labels_frame, height=5, bg=theme["input_bg"], fg=theme["input_fg"])
        labels_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Cadre pour l'aperçu
        preview_frame = tk.LabelFrame(chart_window, text="Aperçu", bg=theme["bg"], fg=theme["fg"])
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Cadre pour les boutons
        button_frame = tk.Frame(chart_window, bg=theme["bg"])
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Variables pour stocker la figure actuelle
        current_figure = {"fig": None}
        
        def parse_data():
            """Analyse les données entrées par l'utilisateur"""
            data_str = data_text.get("1.0", END).strip()
            chart_type = chart_type_var.get()
            
            if not data_str:
                messagebox.showerror("Erreur", "Veuillez entrer des données.")
                return None
            
            try:
                if chart_type == "pie":
                    # Pour les camemberts, on a besoin d'étiquettes et de valeurs
                    values = [float(line.strip()) for line in data_str.split("\n") if line.strip()]
                    labels_str = labels_text.get("1.0", END).strip()
                    
                    if not labels_str:
                        # Générer des étiquettes par défaut
                        labels = [f"Segment {i+1}" for i in range(len(values))]
                    else:
                        labels = [line.strip() for line in labels_str.split("\n") if line.strip()]
                        
                        # S'assurer qu'il y a autant d'étiquettes que de valeurs
                        if len(labels) < len(values):
                            labels.extend([f"Segment {i+1}" for i in range(len(labels), len(values))])
                        elif len(labels) > len(values):
                            labels = labels[:len(values)]
                    
                    return {"labels": labels, "values": values}
                
                elif chart_type in ["line", "scatter"]:
                    # Pour les graphiques XY, on peut avoir des paires x,y
                    lines = [line.strip() for line in data_str.split("\n") if line.strip()]
                    
                    # Vérifier si les données sont au format x,y
                    if "," in lines[0]:
                        x_values = []
                        y_values = []
                        
                        for line in lines:
                            parts = line.split(",")
                            if len(parts) >= 2:
                                try:
                                    x = float(parts[0].strip())
                                    y = float(parts[1].strip())
                                    x_values.append(x)
                                    y_values.append(y)
                                except ValueError:
                                    continue
                        
                        return {"x": x_values, "y": y_values}
                    else:
                        # Sinon, utiliser les indices comme x
                        try:
                            y_values = [float(line) for line in lines]
                            return {"y": y_values}
                        except ValueError:
                            messagebox.showerror("Erreur", "Format de données invalide.")
                            return None
                
                elif chart_type == "bar":
                    # Pour les barres, on peut avoir des paires catégorie,valeur
                    lines = [line.strip() for line in data_str.split("\n") if line.strip()]
                    
                    # Vérifier si les données sont au format catégorie,valeur
                    if "," in lines[0]:
                        categories = []
                        values = []
                        
                        for line in lines:
                            parts = line.split(",")
                            if len(parts) >= 2:
                                try:
                                    cat = parts[0].strip()
                                    val = float(parts[1].strip())
                                    categories.append(cat)
                                    values.append(val)
                                except ValueError:
                                    continue
                        
                        return {"x": categories, "y": values}
                    else:
                        # Sinon, utiliser les indices comme catégories
                        try:
                            values = [float(line) for line in lines]
                            return {"y": values}
                        except ValueError:
                            messagebox.showerror("Erreur", "Format de données invalide.")
                            return None
                
                elif chart_type == "histogram":
                    # Pour l'histogramme, on a juste besoin d'une liste de valeurs
                    try:
                        values = [float(line.strip()) for line in data_str.split("\n") if line.strip()]
                        return {"values": values}
                    except ValueError:
                        messagebox.showerror("Erreur", "Format de données invalide.")
                        return None
                
                else:
                    messagebox.showerror("Erreur", "Type de graphique non pris en charge.")
                    return None
                    
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de l'analyse des données: {e}")
                return None
        
        def preview_chart():
            """Génère un aperçu du graphique"""
            data = parse_data()
            if data is None:
                return
            
            chart_type = chart_type_var.get()
            title = title_entry.get()
            xlabel = xlabel_entry.get()
            ylabel = ylabel_entry.get()
            
            try:
                # Effacer l'aperçu précédent
                for widget in preview_frame.winfo_children():
                    widget.destroy()
                
                # Créer le graphique selon le type
                if chart_type == "line":
                    fig = ChartGenerator.create_line_chart(data, title=title, xlabel=xlabel, ylabel=ylabel)
                elif chart_type == "bar":
                    fig = ChartGenerator.create_bar_chart(data, title=title, xlabel=xlabel, ylabel=ylabel)
                elif chart_type == "pie":
                    fig = ChartGenerator.create_pie_chart(data, title=title)
                elif chart_type == "scatter":
                    fig = ChartGenerator.create_scatter_plot(data, title=title, xlabel=xlabel, ylabel=ylabel)
                elif chart_type == "histogram":
                    fig = ChartGenerator.create_histogram(data, title=title, xlabel=xlabel, ylabel=ylabel)
                else:
                    messagebox.showerror("Erreur", "Type de graphique non pris en charge.")
                    return
                
                # Stocker la figure actuelle
                current_figure["fig"] = fig
                
                # Convertir la figure en image pour l'affichage
                img_data = ChartGenerator.figure_to_image(fig)
                
                # Afficher l'image
                if Image and ImageTk:
                    img = Image.open(img_data)
                    img.thumbnail((600, 400))  # Redimensionner pour l'affichage
                    photo = ImageTk.PhotoImage(img)
                    
                    img_label = Label(preview_frame, image=photo, bg=theme["bg"])
                    img_label.image = photo  # Garder une référence
                    img_label.pack(padx=10, pady=10)
                else:
                    tk.Label(preview_frame, text="Aperçu non disponible (PIL non installé)", bg=theme["bg"], fg=theme["fg"]).pack(padx=10, pady=10)
                
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la génération du graphique: {e}")
        
        def export_to_excel():
            """Exporte le graphique actuel vers Excel"""
            if current_figure["fig"] is None:
                messagebox.showerror("Erreur", "Veuillez d'abord générer un graphique.")
                return
            
            # Demander où sauvegarder le fichier
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Fichiers Excel", "*.xlsx")],
                title="Enregistrer le graphique dans Excel"
            )
            
            if not file_path:
                return  # L'utilisateur a annulé
            
            try:
                # Créer un fichier Excel avec le graphique
                chart_name = title_entry.get() or "Graphique"
                ChartGenerator.create_excel_with_charts(
                    [(current_figure["fig"], chart_name)],
                    file_path
                )
                
                messagebox.showinfo(
                    "Exportation réussie",
                    f"Le graphique a été exporté avec succès dans {file_path}"
                )
            except Exception as e:
                messagebox.showerror(
                    "Erreur d'exportation",
                    f"Erreur lors de l'exportation du graphique: {e}"
                )
        
        # Boutons
        preview_button = Button(
            button_frame,
            text="Aperçu",
            command=preview_chart,
            bg=theme["button_bg"],
            fg=theme["button_fg"],
            width=15
        )
        preview_button.pack(side=tk.LEFT, padx=5)
        
        export_button = Button(
            button_frame,
            text="Exporter vers Excel",
            command=export_to_excel,
            bg=theme["button_bg"],
            fg=theme["button_fg"],
            width=20
        )
        export_button.pack(side=tk.LEFT, padx=5)
        
        close_button = Button(
            button_frame,
            text="Fermer",
            command=chart_window.destroy,
            bg=theme["button_bg"],
            fg=theme["button_fg"],
            width=10
        )
        close_button.pack(side=tk.RIGHT, padx=5)
        
        # Centrer la fenêtre
        chart_window.update_idletasks()
        width = chart_window.winfo_width()
        height = chart_window.winfo_height()
        x = (chart_window.winfo_screenwidth() // 2) - (width // 2)
        y = (chart_window.winfo_screenheight() // 2) - (height // 2)
        chart_window.geometry(f"{width}x{height}+{x}+{y}")

    def generate_sample_charts(self):
        """Génère des exemples de graphiques et les exporte dans un fichier Excel"""
        if plt is None or openpyxl is None:
            messagebox.showerror(
                "Erreur",
                "Impossible de créer des graphiques. Assurez-vous que matplotlib et openpyxl sont installés."
            )
            return
        
        try:
            # Générer des données d'exemple
            sample_data = ChartGenerator.generate_sample_data()
            
            # Créer les graphiques
            line_chart = ChartGenerator.create_line_chart(
                sample_data['line']['multi'],
                title="Exemple de graphique en courbes",
                xlabel="Axe X",
                ylabel="Valeurs"
            )
            
            bar_chart = ChartGenerator.create_bar_chart(
                sample_data['bar'],
                title="Exemple de graphique à barres",
                xlabel="Catégories",
                ylabel="Valeurs"
            )
            
            pie_chart = ChartGenerator.create_pie_chart(
                sample_data['pie'],
                title="Exemple de graphique en camembert"
            )
            
            scatter_chart = ChartGenerator.create_scatter_plot(
                sample_data['scatter'],
                title="Exemple de nuage de points",
                xlabel="X",
                ylabel="Y"
            )
            
            histogram_chart = ChartGenerator.create_histogram(
                sample_data['histogram'],
                title="Exemple d'histogramme",
                xlabel="Valeurs",
                ylabel="Fréquence"
            )
            
            # Demander où sauvegarder le fichier
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Fichiers Excel", "*.xlsx")],
                title="Enregistrer les exemples de graphiques"
            )
            
            if not file_path:
                return  # L'utilisateur a annulé
            
            # Créer un fichier Excel avec tous les graphiques
            charts_data = [
                (line_chart, "Courbes"),
                (bar_chart, "Barres"),
                (pie_chart, "Camembert"),
                (scatter_chart, "Nuage de points"),
                (histogram_chart, "Histogramme")
            ]
            
            ChartGenerator.create_excel_with_charts(charts_data, file_path)
            
            messagebox.showinfo(
                "Exportation réussie",
                f"Les exemples de graphiques ont été exportés avec succès dans {file_path}"
            )
        except Exception as e:
            messagebox.showerror(
                "Erreur",
                f"Erreur lors de la génération des exemples de graphiques: {e}"
            )

    # ----- Fonctions pour changer de mode -----

    def switch_to_local_mode(self, local_api_url=None, local_model=None):
        if local_api_url is None:
            local_api_url = self.config.get(
                "local_api_url", Config.DEFAULT_LOCAL_API_URL
            )
        if local_model is None:
            local_model = self.config.get("local_model", Config.DEFAULT_LOCAL_MODEL)
        if self.config.get("cpu_only"):
            local_api_url += "?cpu_only=true"
        if not is_port_open(local_api_url):
            messagebox.showerror(
                "Erreur",
                "Le serveur local ne semble pas actif. Vérifiez que Ollama serve est lancé.",
            )
            return
        self.config["local_mode"] = True
        self.config["api_url"] = local_api_url
        self.config["model"] = local_model
        # Créer un dictionnaire de configuration du proxy
        proxy_config = {
            "proxy_enabled": self.config.get("proxy_enabled", False),
            "proxy_host": self.config.get("proxy_host", ""),
            "proxy_port": self.config.get("proxy_port", ""),
            "proxy_username": self.config.get("proxy_username", ""),
            "proxy_password": self.config.get("proxy_password", ""),
        }
    
        self.ai_client = AIClient(
            self.config["api_key"],
            self.config["api_url"],
            self.config["model"],
            self.config["timeout"],
            proxy_config
        )
        messagebox.showinfo(
            "Mode local",
            f"Mode local activé.\nAPI : {local_api_url}\nModèle : {local_model}",
        )
        self.update_mode_indicator()

    def switch_to_remote_mode(self):
        self.config["local_mode"] = False
        self.config["eden_mode"] = False  # Désactiver explicitement le mode Eden
        self.config["api_url"] = Config.DEFAULT_API_URL
        self.config["model"] = Config.DEFAULT_MODEL
        
        # Créer un dictionnaire de configuration du proxy
        proxy_config = {
            "proxy_enabled": self.config.get("proxy_enabled", False),
            "proxy_host": self.config.get("proxy_host", ""),
            "proxy_port": self.config.get("proxy_port", ""),
            "proxy_username": self.config.get("proxy_username", ""),
            "proxy_password": self.config.get("proxy_password", ""),
        }
    
        self.ai_client = AIClient(
            self.config["api_key"],
            self.config["api_url"],
            self.config["model"],
            self.config["timeout"],
            proxy_config
        )
        
        # Sauvegarder la configuration mise à jour
        Config.save(self.config)
        
        messagebox.showinfo(
            "Mode distant", "Retour en mode distant (API en ligne) activé."
        )
        self.update_mode_indicator()

    def copy_email_to_clipboard(self, text):
        """Copie le texte dans le presse-papiers et affiche un message de confirmation"""
        try:
            pyperclip.copy(text)
            messagebox.showinfo("Copie réussie", "Le contenu de l'email a été copié dans le presse-papiers.")
            return True
        except Exception as e:
            logging.error(f"Erreur lors de la copie dans le presse-papiers: {e}")
            messagebox.showerror("Erreur", f"Impossible de copier l'email: {e}")
            return False

    def show_email_copy_window(self, email_content):
        """Affiche une fenêtre pour copier le contenu d'un email"""
        email_window = Toplevel(self.root)
        email_window.title("Contenu de l'email")
        email_window.geometry("600x400")
        email_window.resizable(True, True)
        
        # Récupérer le thème actuel
        theme = Config.THEMES.get(self.config['theme'], Config.THEMES['light'])
        
        # Appliquer le thème à la fenêtre
        email_window.config(bg=theme['bg'])
        
        # Zone de texte pour l'affichage de l'email
        email_text = Text(email_window, wrap=tk.WORD, font=("Arial", 10), bg=theme['input_bg'], fg=theme['input_fg'])
        email_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        email_text.insert(END, email_content)
        
        # Rendre le texte sélectionnable mais non modifiable
        email_text.config(state=tk.NORMAL)
        
        # Frame pour les boutons
        button_frame = tk.Frame(email_window, bg=theme['bg'])
        button_frame.pack(padx=10, pady=10)
        
        # Bouton pour copier l'email
        copy_button = Button(
            button_frame,
            text="Copier l'email",
            command=lambda: self.copy_email_to_clipboard(email_content),
            bg=theme['button_bg'],
            fg=theme['button_fg'],
            width=15
        )
        copy_button.pack(side=tk.LEFT, padx=5)
        
        # Bouton pour fermer la fenêtre
        close_button = Button(
            button_frame,
            text="Fermer",
            command=email_window.destroy,
            bg=theme['button_bg'],
            fg=theme['button_fg'],
            width=10
        )
        close_button.pack(side=tk.LEFT, padx=5)

    def detect_and_copy_email(self):
        """Détecte et copie un email avec une reconnaissance très flexible"""
        chat_content = self.chat_box.get("1.0", END)
        
        # Motifs de recherche pour différents formats d'emails
        email_patterns = [
            r'```email\n([\s\S]*?)\n```',
            r'((?:Objet|Sujet|Subject)\s*:.*\n(?:(?:À|To|Destinataire)\s*:.*\n)?(?:(?:De|From|Expéditeur)\s*:.*\n)?\n?[\s\S]+?(?:\n\n|$))',
            r'((?:Bonjour|Hello|Salut).*?(?:Cordialement|Bien cordialement|Salutations)[\s\S]*)',
            r'([\s\S]{20,500}(?:Cordialement|Salutations|Bien à vous)[\s\S]*)'
        ]
        
        for pattern in email_patterns:
            email_match = re.search(pattern, chat_content, re.MULTILINE | re.IGNORECASE | re.DOTALL)
            if email_match:
                email_content = email_match.group(1).strip()
                cleaned_email = re.sub(r'^>\s*', '', email_content, flags=re.MULTILINE)
                cleaned_email = re.sub(r'\n>\s*', '\n', cleaned_email)
                self.show_email_copy_window(cleaned_email)
                return
        
        messagebox.showinfo("Information", "Aucun contenu d'email détecté dans le chat.")
        pass

    def detect_email_in_response(self, response_text):
        """Détecte automatiquement si la réponse contient un email et affiche le bouton pour le copier"""
        # Motifs de recherche pour différents formats d'emails
        email_patterns = [
            r'```email\n([\s\S]*?)\n```',
            r'((?:Objet|Sujet|Subject)\s*:.*\n(?:(?:À|To|Destinataire)\s*:.*\n)?(?:(?:De|From|Expéditeur)\s*:.*\n)?\n?[\s\S]+?(?:\n\n|$))',
            r'((?:Bonjour|Hello|Salut).*?(?:Cordialement|Bien cordialement|Salutations)[\s\S]*)'
            r'([\s\S]{20,500}(?:Cordialement|Salutations|Bien à vous)[\s\S]*)'
        ]
        
        for pattern in email_patterns:
            email_match = re.search(pattern, response_text, re.MULTILINE | re.IGNORECASE | re.DOTALL)
            if email_match:
                # Email détecté, afficher le bouton de copie
                self.add_email_copy_button()
                return True
        
        return False

    def display_assistant_response(self, assistant_response):
        assistant_response = filter_unicode_chars(assistant_response)

        self.append_message("assistant", assistant_response)
        
        # Vérifier s'il y a du code dans la réponse
        lang, code = CodeHandler.extract_code(assistant_response)
        if code:
            self.process_code_in_response(lang, code)
        
        # Détecter automatiquement si un email est présent
        self.detect_email_in_response(assistant_response)
        self.detect_excel_processing_request(assistant_response)
        self.history_manager.add_message("assistant", assistant_response)
        self.stop_loading()
        
        # Effacer les pièces jointes après l'affichage de la réponse
        for widget in self.attachments_frame.winfo_children():
            widget.destroy()
        self.attachments.clear()
        
    def switch_to_eden_mode(self):
        """Bascule l'application en mode Eden AI"""
        self.config["eden_mode"] = True
        self.config["local_mode"] = False

        
        # Créer un dictionnaire de configuration du proxy
        proxy_config = {
            "proxy_enabled": self.config.get("proxy_enabled", False),
            "proxy_host": self.config.get("proxy_host", ""),
            "proxy_port": self.config.get("proxy_port", ""),
            "proxy_username": self.config.get("proxy_username", ""),
            "proxy_password": self.config.get("proxy_password", ""),
        }
        
        # Initialiser avec les paramètres corrects depuis la configuration
        eden_api_url = self.config.get("eden_api_url", Config.DEFAULT_EDEN_API_URL)
        eden_api_key = self.config.get("eden_api_key", Config.DEFAULT_EDEN_API_KEY)
        eden_geni_id = self.config.get("eden_geni_id", Config.DEFAULT_EDEN_GENI_ID)
        
        # Créer une nouvelle instance du client Eden avec les proxies
        self.ai_client = EdenClient(self.config["timeout"], proxy_config)
        
        # Mettre à jour les paramètres si nécessaire
        self.ai_client.api_url = eden_api_url
        self.ai_client.api_key = eden_api_key
        self.ai_client.geni_id = eden_geni_id
        
        # Sauvegarder les modifications de configuration
        Config.save(self.config)
        
        messagebox.showinfo(
            "Mode Eden",
            "Mode Eden AI activé.\nUtilisation de l'API Eden pour les conversations."
        )
        self.update_mode_indicator()

# ------------------- Fonction principale -------------------
def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[logging.FileHandler("app.log"), logging.StreamHandler()],
    )
    root = tk.Tk()
    
    # Vérifier spécifiquement pour openpyxl qui est nécessaire pour les fichiers Excel
    try:
        import openpyxl
        logging.info("Module openpyxl trouvé, version: " + openpyxl.__version__)
    except ImportError:
        logging.warning("Le module openpyxl n'est pas installé. L'analyse des fichiers Excel sera limitée.")
        messagebox.showwarning(
            "Dépendance manquante", 
            "Le module 'openpyxl' n'est pas installé.\n\n"
            "Les fichiers Excel ne pourront pas être analysés en détail.\n"
            "Pour installer openpyxl, exécutez la commande :\n"
            "pip install openpyxl"
        )
    
    app = ChatApplication(root)
    root.mainloop()


if __name__ == "__main__":
    main()
    
            

class SystemPreflightCheck:
    # Liste des modules requis avec leurs versions minimales recommandées
    REQUIRED_MODULES = {
        'tkinter': None,  # Tkinter est généralement inclus avec Python
        'requests': '2.25.0',
        'pygments': '2.7.0',
        'pyperclip': '1.8.0',
        'pillow': '8.0.0',
        'matplotlib': '3.3.0',
        'numpy': '1.19.0',
        'pandas': '1.1.0',
        'openpyxl': '3.0.0',
        'pytesseract': '0.3.8',
        'opencv-python': '4.5.0',
    }

    @staticmethod
    def check_dependencies(show_dialog=True):
        """Vérifie toutes les dépendances requises pour l'application"""
        missing_modules = []
        outdated_modules = []
        installed_modules = []
        
        for module_name, min_version in SystemPreflightCheck.REQUIRED_MODULES.items():
            try:
                if module_name == 'Pillow':
                    module = importlib.import_module('PIL')
                else:
                    module = importlib.import_module(module_name)
                
                # Vérifier la version si spécifiée
                if hasattr(module, '__version__') and min_version:
                    current_version = module.__version__
                    if current_version < min_version:
                        outdated_modules.append((module_name, current_version, min_version))
                    else:
                        installed_modules.append((module_name, current_version))
                else:
                    installed_modules.append((module_name, "Version non spécifiée"))
                    
            except ImportError:
                missing_modules.append(module_name)
        
        if show_dialog and (missing_modules or outdated_modules):
            SystemPreflightCheck.show_dependencies_dialog(
                missing_modules, 
                outdated_modules,
                installed_modules
            )
            
        return missing_modules, outdated_modules, installed_modules

    @staticmethod
    def show_dependencies_dialog(missing_modules, outdated_modules, installed_modules):
        """Affiche une fenêtre de dialogue avec l'état des dépendances"""
        dialog = tk.Toplevel()
        dialog.title("Vérification des Dépendances Python")
        dialog.geometry("600x400")
        dialog.resizable(True, True)
        
        # Configuration du style
        style = ttk.Style()
        style.configure("Missing.TLabel", foreground="red")
        style.configure("Outdated.TLabel", foreground="orange")
        style.configure("Installed.TLabel", foreground="green")
        
        # Frame principal avec scrollbar
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Création du canvas et scrollbar
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Titre
        ttk.Label(
            scrollable_frame,
            text="État des Dépendances",
            font=("Arial", 12, "bold")
        ).pack(pady=10)
        
        # Modules manquants
        if missing_modules:
            ttk.Label(
                scrollable_frame,
                text="Modules manquants :",
                font=("Arial", 10, "bold")
            ).pack(anchor="w", pady=5)
            for module in missing_modules:
                frame = ttk.Frame(scrollable_frame)
                frame.pack(fill="x", padx=20)
                ttk.Label(
                    frame,
                    text=f"• {module}",
                    style="Missing.TLabel"
                ).pack(side="left")
                ttk.Button(
                    frame,
                    text="Installer",
                    command=lambda m=module: SystemPreflightCheck.install_module(m, dialog)
                ).pack(side="right")
        
        # Modules obsolètes
        if outdated_modules:
            ttk.Label(
                scrollable_frame,
                text="Modules à mettre à jour :",
                font=("Arial", 10, "bold")
            ).pack(anchor="w", pady=5)
            for module, current, required in outdated_modules:
                frame = ttk.Frame(scrollable_frame)
                frame.pack(fill="x", padx=20)
                ttk.Label(
                    frame,
                    text=f"• {module} (actuel: {current}, requis: {required})",
                    style="Outdated.TLabel"
                ).pack(side="left")
                ttk.Button(
                    frame,
                    text="Mettre à jour",
                    command=lambda m=module: SystemPreflightCheck.update_module(m, dialog)
                ).pack(side="right")
        
        # Modules installés
        if installed_modules:
            ttk.Label(
                scrollable_frame,
                text="Modules installés :",
                font=("Arial", 10, "bold")
            ).pack(anchor="w", pady=5)
            for module, version in installed_modules:
                ttk.Label(
                    scrollable_frame,
                    text=f"• {module} ({version})",
                    style="Installed.TLabel"
                ).pack(anchor="w", padx=20)
        
        # Boutons d'action
        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill="x", padx=10, pady=10)
        
        if missing_modules or outdated_modules:
            ttk.Button(
                button_frame,
                text="Tout installer/mettre à jour",
                command=lambda: SystemPreflightCheck.install_all_dependencies(dialog)
            ).pack(side="left", padx=5)
        
        ttk.Button(
            button_frame,
            text="Fermer",
            command=dialog.destroy
        ).pack(side="right", padx=5)
        
        # Configuration finale du scrolling
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Centrer la fenêtre
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")

    @staticmethod
    def install_module(module_name, parent_window):
        """Installe un module Python via pip"""
        try:
            import subprocess
            
            # Créer une fenêtre de progression
            progress_window = tk.Toplevel(parent_window)
            progress_window.title(f"Installation de {module_name}")
            progress_window.geometry("300x100")
            
            # Label pour le statut
            status_label = ttk.Label(
                progress_window,
                text=f"Installation de {module_name} en cours..."
            )
            status_label.pack(pady=10)
            
            # Barre de progression indéterminée
            progress = ttk.Progressbar(
                progress_window,
                mode='indeterminate'
            )
            progress.pack(fill='x', padx=20)
            progress.start()
            
            # Fonction pour exécuter l'installation en arrière-plan
            def run_installation():
                try:
                    subprocess.check_call(
                        [sys.executable, "-m", "pip", "install", module_name],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE
                    )
                    progress_window.after(
                        0,
                        lambda: SystemPreflightCheck.installation_complete(
                            progress_window,
                            parent_window,
                            True,
                            module_name
                        )
                    )
                except subprocess.CalledProcessError as e:
                    progress_window.after(
                        0,
                        lambda: SystemPreflightCheck.installation_complete(
                            progress_window,
                            parent_window,
                            False,
                            module_name,
                            str(e)
                        )
                    )
            
            # Lancer l'installation dans un thread séparé
            threading.Thread(target=run_installation, daemon=True).start()
            
        except Exception as e:
            messagebox.showerror(
                "Erreur",
                f"Erreur lors de l'installation de {module_name}: {str(e)}"
            )

    @staticmethod
    def update_module(module_name, parent_window):
        """Met à jour un module Python via pip"""
        SystemPreflightCheck.install_module(f"{module_name} --upgrade", parent_window)

    @staticmethod
    def install_all_dependencies(parent_window):
        """Installe ou met à jour toutes les dépendances manquantes ou obsolètes"""
        missing_modules, outdated_modules, _ = SystemPreflightCheck.check_dependencies(False)
        
        for module in missing_modules:
            SystemPreflightCheck.install_module(module, parent_window)
        
        for module, _, _ in outdated_modules:
            SystemPreflightCheck.update_module(module, parent_window)

    @staticmethod
    def installation_complete(progress_window, parent_window, success, module_name, error_msg=None):
        """Gère la fin de l'installation d'un module"""
        progress_window.destroy()
        
        if success:
            messagebox.showinfo(
                "Installation réussie",
                f"Le module {module_name} a été installé avec succès."
            )
            # Rafraîchir la fenêtre des dépendances
            parent_window.destroy()
            SystemPreflightCheck.check_dependencies()
        else:
            messagebox.showerror(
                "Erreur d'installation",
                f"Erreur lors de l'installation de {module_name}:\n{error_msg}"
            )

    @staticmethod
    def check_permissions():
        """Vérifie les permissions de fichiers et répertoires"""
        check_paths = [
            Config.CONFIG_FILE,
            Config.HISTORY_FILE,
            Config.PERSISTENCE_FILE,
            'app.log'
        ]
        permission_issues = []
        for path in check_paths:
            try:
                with open(path, 'a') as f:
                    pass
            except (PermissionError, IOError) as e:
                permission_issues.append((path, str(e)))
        return permission_issues

    @staticmethod
    def check_network_connectivity():
        """Vérifie la connectivité réseau"""
        try:
            socket.create_connection(('www.google.com', 80), timeout=5)
            return True
        except (socket.error, socket.timeout):
            return False

    @staticmethod
    def preflight_check():
        """Effectue une vérification complète avant le lancement"""
        missing_modules, outdated_modules, _ = SystemPreflightCheck.check_dependencies(True)
        
        if missing_modules:
            error_msg = 'Modules manquants : ' + ', '.join(missing_modules)
            messagebox.showwarning('Dépendances Manquantes', error_msg)
        
        permission_issues = SystemPreflightCheck.check_permissions()
        if permission_issues:
            error_details = '\n'.join([f'{path}: {error}' for path, error in permission_issues])
            messagebox.showerror(
                'Erreurs de Permissions',
                f"Problèmes d'accès aux fichiers :\n{error_details}"
            )
            return False
        
        if not Config.load().get('local_mode', False):
            if not SystemPreflightCheck.check_network_connectivity():
                messagebox.showwarning(
                    'Connectivité Réseau',
                    'Pas de connexion internet détectée. Le mode local sera utilisé.'
                )
        
        return len(missing_modules) == 0 and len(permission_issues) == 0
